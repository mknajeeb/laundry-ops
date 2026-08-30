"""Payroll analytics dashboard: period comparison, grouping, PDF page-1, reconciliations."""

from __future__ import annotations

from backend.payroll_report import (
    build_payroll_report_html,
    build_payroll_report_xlsx,
    build_report_row,
)
from backend.payroll_report_analytics import (
    aggregate_period_metrics,
    build_kpi_cards,
    build_report_analytics,
    group_rows_by_period_then_pay_date,
    select_comparison_periods,
)


def _row(
    *,
    name,
    cat,
    ps,
    pe,
    pay_date,
    reg,
    ot,
    rate=17.0,
    ot_rate=25.5,
    line_id=1,
    batch_id=1,
    user_id=1,
    paid_full=True,
    amount_paid=None,
    employer_taxes=0.0,
    withheld=0.0,
):
    total_hours = float(reg) + float(ot)
    base = round(total_hours * rate, 2)
    premium = round(float(ot) * (ot_rate - rate), 2)
    gross = round(base + premium, 2)
    paid = amount_paid if amount_paid is not None else gross
    line = {
        "id": line_id,
        "user_id": user_id,
        "worker_name_snapshot": name,
        "approved_hours": reg,
        "ot_hours": ot,
        "rate": rate,
        "ot_rate": ot_rate,
        "gross_amount": gross,
        "total_amount": gross,
        "gross_wages": gross,
        "sick_pay_amount": 0,
        "bonus_tip_amount": 0,
        "reimbursement_amount": 0,
        "adjustments": 0,
        "payment_status": "paid",
        "net_pay": gross - withheld,
        "payout_details": {
            "settlement": {
                "amount_paid": paid,
                "amount_withheld": withheld,
                "outstanding_balance": round(max(0.0, gross - paid), 2),
                "paid_full_gross_without_withholding": paid_full,
            },
            "employer_taxes": {
                "employer_social_security": employer_taxes,
            },
        },
    }
    batch = {
        "id": batch_id,
        "batch_name": f"{cat}-{ps}",
        "worker_category": cat,
        "pay_period_start": ps,
        "pay_period_end": pe,
        "status": "paid",
        "official_pay_date": pay_date,
        "payout_details_finalized_at": "2026-06-01T00:00:00",
    }
    return build_report_row(batch, line, report_type="monthly_paid")


def test_select_comparison_periods_includes_selected_and_prior():
    all_p = [
        ("2026-05-04", "2026-05-10"),
        ("2026-05-11", "2026-05-17"),
        ("2026-05-18", "2026-05-24"),
        ("2026-05-25", "2026-05-31"),
        ("2026-06-01", "2026-06-07"),
        ("2026-06-08", "2026-06-14"),
    ]
    selected = select_comparison_periods(
        all_p,
        anchor_periods=[("2026-06-08", "2026-06-14")],
        comparison_range=4,
    )
    assert len(selected) == 4
    assert selected[-1] == ("2026-06-08", "2026-06-14")
    assert selected == all_p[-4:]
    assert selected == sorted(selected)


def test_monthly_paid_comparison_excludes_prior_month_pay_date_periods(monkeypatch):
    """July Monthly Paid charts months, not June-paid payroll weeks."""
    july_rows = [
        _row(
            name="A",
            cat="w2",
            ps="2026-06-22",
            pe="2026-06-28",
            pay_date="2026-07-04",
            reg=40,
            ot=0,
            line_id=1,
            user_id=1,
        ),
        _row(
            name="B",
            cat="temp",
            ps="2026-06-22",
            pe="2026-06-28",
            pay_date="2026-07-04",
            reg=40,
            ot=0,
            line_id=2,
            user_id=2,
        ),
        _row(
            name="C",
            cat="contractor_1099",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=2,
            line_id=3,
            user_id=3,
        ),
    ]

    def _fake_monthly(conn, oid, *, month, year, **kwargs):
        if (year, month) == (2026, 6):
            return [
                _row(
                    name="J",
                    cat="w2",
                    ps="2026-06-01",
                    pe="2026-06-07",
                    pay_date="2026-06-13",
                    reg=40,
                    ot=0,
                    line_id=10,
                    user_id=10,
                )
            ]
        if (year, month) == (2026, 7):
            return july_rows
        return []

    monkeypatch.setattr(
        "backend.payroll_report_analytics.list_org_periods_asc",
        lambda *a, **k: [
            ("2026-06-01", "2026-06-07"),
            ("2026-06-22", "2026-06-28"),
            ("2026-07-06", "2026-07-12"),
        ],
    )
    monkeypatch.setattr(
        "backend.payroll_report_analytics._fetch_monthly_paid_rows",
        _fake_monthly,
    )
    monkeypatch.setattr(
        "backend.payroll_report_analytics.fetch_rows_for_periods",
        lambda *a, **k: (_ for _ in ()).throw(
            AssertionError("monthly mode must not fetch period rows for trend")
        ),
    )

    class _Conn:
        def cursor(self, dictionary=False):
            raise AssertionError("should not hit DB cursor")

    analytics = build_report_analytics(
        _Conn(),
        3,
        detail_rows=july_rows,
        report_type="monthly_paid",
        filters={"month": 7, "year": 2026, "include_employee_detail": True},
        trend_range=4,
        compare_with="previous_month",
    )
    assert analytics["comparison_mode"] == "month"
    assert analytics["compare_with"] == "previous_month"
    assert analytics["summary"]["focus_period"].startswith("July 2026")
    assert analytics["summary"]["previous_period"] == "June 2026"
    months = [e["month"] for e in analytics["month_comparison"]]
    assert "2026-07" in months
    assert all(len(m) == 7 for m in months)  # YYYY-MM buckets
    # Top-level period_comparison stays empty in month mode; periods nest under months.
    assert analytics["period_comparison"] == []
    july = next(e for e in analytics["month_comparison"] if e["month"] == "2026-07")
    assert july["worker_count"] == 3  # distinct user ids
    assert july["avg_cost_per_hour"] == round(
        july["total_payroll_cost"] / july["total_hours"], 2
    )
    assert july["pay_dates_label"]
    periods = july.get("periods") or []
    assert len(periods) == 2
    assert [p["pay_period_start"] for p in periods] == ["2026-06-22", "2026-07-06"]
    assert periods[0]["pay_dates_label"] == "2026-07-04"
    assert periods[1]["pay_dates_label"] == "2026-07-18"
    assert periods[0]["worker_count"] == 2
    assert periods[1]["worker_count"] == 1
    assert analytics["executive_narrative"]["headline"]
    assert any(c["key"] == "avg_hours_per_worker" for c in analytics["kpis"])


def test_period_trend_options_counts():
    from backend.payroll_report_analytics import normalize_trend_range, PERIOD_TREND_OPTIONS

    assert PERIOD_TREND_OPTIONS == (3, 4, 5, 8)
    for n in PERIOD_TREND_OPTIONS:
        assert normalize_trend_range(n, mode="period") == n


def test_same_month_last_year_compare(monkeypatch):
    rows = [
        _row(
            name="A",
            cat="w2",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=0,
            line_id=1,
            user_id=1,
        )
    ]

    def _fake_monthly(conn, oid, *, month, year, **kwargs):
        if (year, month) == (2025, 7):
            return [
                _row(
                    name="Old",
                    cat="w2",
                    ps="2025-07-07",
                    pe="2025-07-13",
                    pay_date="2025-07-19",
                    reg=20,
                    ot=0,
                    line_id=9,
                    user_id=9,
                )
            ]
        return []

    monkeypatch.setattr(
        "backend.payroll_report_analytics.list_org_periods_asc", lambda *a, **k: []
    )
    monkeypatch.setattr(
        "backend.payroll_report_analytics._fetch_monthly_paid_rows", _fake_monthly
    )

    class _Conn:
        def cursor(self, dictionary=False):
            raise AssertionError("no cursor")

    analytics = build_report_analytics(
        _Conn(),
        3,
        detail_rows=rows,
        report_type="monthly_paid",
        filters={"month": 7, "year": 2026},
        compare_with="same_month_last_year",
        trend_range=3,
    )
    assert analytics["compare_with"] == "same_month_last_year"
    assert analytics["summary"]["previous_period"] == "July 2025"
    assert analytics["kpis"][0]["previous"] is not None


def test_monthly_groups_period_then_pay_date():
    rows = [
        _row(
            name="Maria Perez",
            cat="contractor_1099",
            ps="2026-06-08",
            pe="2026-06-14",
            pay_date="2026-06-20",
            reg=40,
            ot=4.65,
            line_id=1,
            user_id=10,
        ),
        _row(
            name="Guiying Lin",
            cat="temp",
            ps="2026-06-08",
            pe="2026-06-14",
            pay_date="2026-06-20",
            reg=40,
            ot=7.4,
            line_id=2,
            user_id=11,
        ),
        _row(
            name="Other",
            cat="w2",
            ps="2026-06-01",
            pe="2026-06-07",
            pay_date="2026-06-13",
            reg=40,
            ot=0,
            line_id=3,
            user_id=12,
            employer_taxes=50,
            withheld=100,
            paid_full=False,
            amount_paid=580,
        ),
    ]
    groups = group_rows_by_period_then_pay_date(rows)
    assert [g["payroll_period"] for g in groups] == [
        "2026-06-01 – 2026-06-07",
        "2026-06-08 – 2026-06-14",
    ]
    assert groups[1]["pay_dates"][0]["pay_date"] == "2026-06-20"
    names = [r["employee_name"] for r in groups[1]["pay_dates"][0]["rows"]]
    assert names == ["Guiying Lin", "Maria Perez"]


def test_maria_guiying_ot_gross_definitions():
    maria = _row(
        name="Maria Perez",
        cat="contractor_1099",
        ps="2026-06-22",
        pe="2026-06-28",
        pay_date="2026-07-04",
        reg=40,
        ot=4.65,
        line_id=288,
        user_id=20,
    )
    assert abs(maria["regular_hours"] + maria["ot_hours"] - 44.65) < 0.01
    assert abs(maria["base_earnings"] - 759.05) < 0.02
    assert abs(maria["ot_premium"] - 39.53) < 0.02
    assert abs(maria["gross_pay"] - 798.58) < 0.02
    assert abs(maria["net_pay"] - maria["gross_pay"]) < 0.01
    assert abs(maria["total_payroll_cost"] - maria["gross_pay"]) < 0.01

    guiying = _row(
        name="Guiying Lin",
        cat="temp",
        ps="2026-06-08",
        pe="2026-06-14",
        pay_date="2026-06-20",
        reg=40,
        ot=7.4,
        line_id=203,
        user_id=21,
    )
    assert abs(guiying["regular_hours"] + guiying["ot_hours"] - 47.40) < 0.01
    assert abs(guiying["base_earnings"] - 805.80) < 0.02
    assert abs(guiying["ot_premium"] - 62.90) < 0.02
    assert abs(guiying["gross_pay"] - 868.70) < 0.02
    assert abs(guiying["net_pay"] - 868.70) < 0.02
    assert abs(guiying["total_payroll_cost"] - 868.70) < 0.02


def test_employer_taxes_added_once_employee_taxes_not_in_cost():
    row = _row(
        name="W2 Worker",
        cat="w2",
        ps="2026-06-01",
        pe="2026-06-07",
        pay_date="2026-06-13",
        reg=40,
        ot=0,
        rate=20,
        ot_rate=30,
        employer_taxes=62.0,
        withheld=120.0,
        paid_full=False,
        amount_paid=680.0,
    )
    assert abs(row["gross_pay"] - 800.0) < 0.01
    assert abs(row["employer_taxes"] - 62.0) < 0.01
    assert abs(row["total_payroll_cost"] - 862.0) < 0.01
    # EE taxes must not inflate total payroll cost
    assert abs(row["total_payroll_cost"] - (row["gross_pay"] + row["employer_taxes"])) < 0.01


def test_dashboard_totals_equal_detail_rows():
    rows = [
        _row(
            name="A",
            cat="temp",
            ps="2026-06-08",
            pe="2026-06-14",
            pay_date="2026-06-20",
            reg=40,
            ot=2,
            line_id=1,
            user_id=1,
        ),
        _row(
            name="B",
            cat="contractor_1099",
            ps="2026-06-08",
            pe="2026-06-14",
            pay_date="2026-06-20",
            reg=30,
            ot=0,
            line_id=2,
            user_id=2,
        ),
    ]
    metrics = aggregate_period_metrics(rows)
    assert abs(metrics["gross_pay"] - sum(r["gross_pay"] for r in rows)) < 0.01
    assert abs(metrics["total_payroll_cost"] - sum(r["total_payroll_cost"] for r in rows)) < 0.01
    assert metrics["worker_count"] == 2


def test_kpi_neutral_trend_for_cost_increase():
    current = {
        "total_hours": 100,
        "regular_hours": 90,
        "ot_hours": 10,
        "gross_pay": 2000,
        "ot_premium": 85,
        "employee_tax_deductions": 0,
        "net_pay": 2000,
        "employer_taxes": 50,
        "total_payroll_cost": 2050,
        "amount_paid": 2000,
        "outstanding_balance": 0,
        "worker_count": 5,
        "avg_cost_per_hour": 20.5,
        "avg_pay_rate": 20.0,
        "avg_hours_per_worker": 20.0,
        "ot_pct_of_hours": 10.0,
    }
    previous = {
        **current,
        "total_payroll_cost": 1800,
        "ot_hours": 5,
        "ot_premium": 40,
        "avg_cost_per_hour": 18.0,
        "avg_pay_rate": 18.0,
        "avg_hours_per_worker": 18.0,
        "ot_pct_of_hours": 5.0,
    }
    cards = {c["key"]: c for c in build_kpi_cards(current, previous)}
    assert set(cards) == {
        "total_payroll_cost",
        "gross_pay",
        "total_hours",
        "worker_count",
        "avg_hours_per_worker",
        "avg_cost_per_hour",
    }
    assert [c["key"] for c in build_kpi_cards(current, previous)] == [
        "total_payroll_cost",
        "gross_pay",
        "total_hours",
        "worker_count",
        "avg_hours_per_worker",
        "avg_cost_per_hour",
    ]
    assert cards["avg_cost_per_hour"]["label"] == "Average Employer Cost / Hour"
    assert "amount_paid" not in cards
    assert "outstanding_balance" not in cards
    assert cards["total_payroll_cost"]["direction"] == "up"
    assert cards["total_payroll_cost"]["neutral_trend"] is True
    assert cards["total_payroll_cost"]["previous"] == 1800
    assert cards["total_payroll_cost"]["diff"] == 250
    # Absolute hours change only — no % for avg hours/worker.
    assert cards["avg_hours_per_worker"]["diff"] == 2.0
    assert cards["avg_hours_per_worker"]["pct"] is None
    from backend.payroll_report_analytics import build_ot_summary

    narrative = __import__(
        "backend.payroll_report_analytics", fromlist=["build_executive_narrative"]
    ).build_executive_narrative(current, previous, [])
    assert all("Hours/worker" not in d for d in narrative["drivers"])

    ot = build_ot_summary(current, previous)
    assert ot["value"] == 10
    assert ot["ot_pct_of_hours"] == 10.0
    assert ot["direction"] == "up"
    assert ot["ot_premium"] == 85
    assert ot["ot_premium_pct_of_gross"] == round(85 / 2000 * 100, 2)
    assert ot["previous_ot_hours"] == 5
    assert ot["previous_ot_premium"] == 40
    assert ot["ot_premium_diff"] == 45


def test_avg_cost_and_pay_rate_definitions():
    rows = [
        _row(
            name="A",
            cat="temp",
            ps="2026-06-08",
            pe="2026-06-14",
            pay_date="2026-06-20",
            reg=40,
            ot=2,
            line_id=1,
            user_id=1,
        ),
    ]
    metrics = aggregate_period_metrics(rows)
    assert metrics["avg_cost_per_hour"] == round(
        metrics["total_payroll_cost"] / metrics["total_hours"], 2
    )
    assert metrics["avg_pay_rate"] == round(metrics["gross_pay"] / metrics["total_hours"], 2)
    # Avg pay rate includes OT premium in gross.
    assert metrics["gross_pay"] == round(
        metrics["base_earnings"] + metrics["ot_premium"] + metrics["other_earnings"], 2
    )


def test_category_base_plus_ot_premium_equals_gross_all_categories():
    from backend.payroll_report_analytics import category_breakdown

    rows = [
        _row(
            name="W2 Worker",
            cat="w2",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=8.65,
            rate=17.0,
            ot_rate=25.5,
            line_id=1,
            user_id=1,
            employer_taxes=100.0,
            withheld=50.0,
        ),
        _row(
            name="Temp Worker",
            cat="temp",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=12.49,
            rate=17.0,
            ot_rate=25.5,
            line_id=2,
            user_id=2,
        ),
        _row(
            name="1099 Worker",
            cat="contractor_1099",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=7.02,
            rate=17.0,
            ot_rate=25.5,
            line_id=3,
            user_id=3,
        ),
    ]
    rows[0]["total_payroll_cost"] = round(rows[0]["gross_pay"] + rows[0]["employer_taxes"], 2)

    cats = {c["worker_category"]: c for c in category_breakdown(rows)}
    for cat_key in ("w2", "temp", "contractor_1099"):
        c = cats[cat_key]
        detail = [r for r in rows if r["worker_category"] == cat_key]
        assert abs(c["base_earnings"] + c["ot_premium"] + c["other_earnings"] - c["gross_pay"]) < 0.005
        assert c["gross_reconciles"] is True
        assert c["gross_reconciliation_diff"] == 0.0
        assert abs(c["base_earnings"] - sum(r["base_earnings"] for r in detail)) < 0.005
        assert abs(c["ot_premium"] - sum(r["ot_premium"] for r in detail)) < 0.005
        assert abs(c["gross_pay"] - sum(r["gross_pay"] for r in detail)) < 0.005

    # Employer taxes raise Cost/Hour but not Avg Pay Rate.
    w2 = cats["w2"]
    assert w2["avg_pay_rate"] == round(w2["gross_pay"] / w2["total_hours"], 2)
    assert w2["avg_cost_per_hour"] == round(w2["total_payroll_cost"] / w2["total_hours"], 2)
    assert w2["avg_cost_per_hour"] > w2["avg_pay_rate"]
    assert cats["temp"]["avg_pay_rate"] == cats["temp"]["avg_cost_per_hour"]
    assert cats["contractor_1099"]["avg_pay_rate"] == cats["contractor_1099"]["avg_cost_per_hour"]


def test_w2_workforce_avg_pay_rate_and_cost_per_hour_production_example():
    """Production Jul 6–12 W-2: $17.32 Avg Pay Rate, $19.45 Cost/Hour."""
    from backend.payroll_report_analytics import category_breakdown, workforce_breakdown_totals

    # Six W-2 workers at $17/hr matching org 3 period 2026-07-06 – 2026-07-12.
    specs = [
        ("Amna Yousaf", 38.16, 0.0, 1),
        ("Evelin Delgado Hernandez", 39.97, 0.0, 2),
        ("Jasanpreet Singh", 24.91, 0.0, 3),
        ("Joshua Cuenca", 35.91, 0.0, 4),
        ("Tarannum Mithila", 39.67, 0.0, 5),
        ("Varun Kumar Mongia", 40.0, 8.65, 6),
    ]
    # Employer tax total $483.30 allocated proportional to gross for the category rollup.
    rows = []
    grosses = []
    for name, reg, ot, uid in specs:
        r = _row(
            name=name,
            cat="w2",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=reg,
            ot=ot,
            rate=17.0,
            ot_rate=25.5,
            line_id=uid,
            user_id=uid,
            employer_taxes=0.0,
            withheld=50.0,
        )
        grosses.append(r["gross_pay"])
        rows.append(r)
    total_gross = round(sum(grosses), 2)
    assert total_gross == 3937.12
    # Distribute ER taxes so category ER sum equals production $483.30.
    er_total = 483.30
    allocated = 0.0
    for i, r in enumerate(rows):
        if i < len(rows) - 1:
            share = round(er_total * (grosses[i] / total_gross), 2)
            allocated = round(allocated + share, 2)
        else:
            share = round(er_total - allocated, 2)
        r["employer_taxes"] = share
        r["total_payroll_cost"] = round(r["gross_pay"] + share, 2)

    cats = category_breakdown(rows)
    w2 = next(c for c in cats if c["worker_category"] == "w2")
    assert w2["total_hours"] == 227.27
    assert w2["ot_hours"] == 8.65
    assert w2["base_earnings"] == 3863.59
    assert w2["ot_premium"] == 73.53
    assert w2["gross_pay"] == 3937.12
    assert abs(w2["base_earnings"] + w2["ot_premium"] - w2["gross_pay"]) < 0.005
    assert abs(w2["employer_taxes"] - 483.30) < 0.02
    assert abs(w2["total_payroll_cost"] - 4420.42) < 0.02
    assert w2["avg_pay_rate"] == 17.32
    assert w2["avg_cost_per_hour"] == 19.45
    # Category gross must equal sum of detail-row gross exactly.
    assert abs(w2["gross_pay"] - sum(r["gross_pay"] for r in rows)) < 0.005

    totals = workforce_breakdown_totals(cats)
    assert totals["avg_pay_rate"] == 17.32
    assert totals["avg_cost_per_hour"] == 19.45
    assert totals["base_earnings"] == 3863.59
    assert totals["ot_premium"] == 73.53


def test_category_gross_equals_sum_of_detail_gross():
    from backend.payroll_report_analytics import category_breakdown

    rows = [
        _row(
            name="A",
            cat="w2",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=0,
            line_id=1,
            user_id=1,
            employer_taxes=100,
            withheld=80,
        ),
        _row(
            name="B",
            cat="w2",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=5,
            line_id=2,
            user_id=2,
            employer_taxes=120,
            withheld=90,
        ),
        _row(
            name="C",
            cat="temp",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=2,
            line_id=3,
            user_id=3,
        ),
    ]
    # Rebuild total_payroll_cost after employer_taxes on W-2 rows.
    for r in rows:
        if r["worker_category"] == "w2":
            r["total_payroll_cost"] = round(r["gross_pay"] + r["employer_taxes"], 2)

    cats = {c["worker_category"]: c for c in category_breakdown(rows)}
    w2_detail = [r for r in rows if r["worker_category"] == "w2"]
    assert cats["w2"]["gross_pay"] == round(sum(r["gross_pay"] for r in w2_detail), 2)
    assert cats["w2"]["base_earnings"] == round(sum(r["base_earnings"] for r in w2_detail), 2)
    assert cats["w2"]["ot_premium"] == round(sum(r["ot_premium"] for r in w2_detail), 2)
    # Temp/1099: Avg Pay Rate == Cost/Hour when no ER tax.
    assert cats["temp"]["avg_pay_rate"] == cats["temp"]["avg_cost_per_hour"]


def test_pdf_dashboard_before_detail_and_contains_charts():
    rows = [
        _row(
            name="Maria Perez",
            cat="contractor_1099",
            ps="2026-06-08",
            pe="2026-06-14",
            pay_date="2026-06-20",
            reg=40,
            ot=4.65,
            line_id=1,
            user_id=1,
        ),
        _row(
            name="Guiying Lin",
            cat="temp",
            ps="2026-06-01",
            pe="2026-06-07",
            pay_date="2026-06-13",
            reg=40,
            ot=7.4,
            line_id=2,
            user_id=2,
        ),
    ]
    # Minimal analytics without DB for PDF structure
    from backend.payroll_report import _sum_totals, _build_summary
    from backend.payroll_report_analytics import (
        build_kpi_cards,
        build_ot_summary,
        build_period_comparison_entries,
        category_breakdown,
        group_rows_by_period_then_pay_date,
        workforce_breakdown_totals,
    )

    groups = group_rows_by_period_then_pay_date(rows)
    by_period = {}
    for r in rows:
        key = (r["pay_period_start"], r["pay_period_end"])
        by_period.setdefault(key, []).append(r)
    ordered = sorted(by_period.keys())
    pc = build_period_comparison_entries(by_period, ordered)
    detail = aggregate_period_metrics(rows)
    cats = category_breakdown(rows)
    prev = pc[0] if len(pc) > 1 else None
    analytics = {
        "summary": {
            **detail,
            "payroll_period_count": len(groups),
            "official_pay_date_count": 2,
            "comparison_range": 4,
        },
        "kpis": build_kpi_cards(detail, prev),
        "ot_summary": build_ot_summary(detail, prev),
        "period_comparison": pc,
        "category_breakdown": cats,
        "workforce_totals": workforce_breakdown_totals(cats),
        "overtime_analysis": pc,
        "groups": groups,
        "comparison_range": 4,
    }
    totals = _sum_totals(rows)
    report = {
        "rows": rows,
        "totals": totals,
        "summary": _build_summary(rows, totals),
        "analytics": analytics,
        "groups": groups,
        "report_type": "monthly_paid",
        "report_heading": "Monthly Payroll Paid: June 2026",
        "filters": {"report_type": "monthly_paid", "month": 6, "year": 2026, "comparison_range": 4},
        "date_match_rule": "test",
    }
    html = build_payroll_report_html(report)
    dash_idx = html.find('id="payroll-analytics-dashboard"')
    detail_idx = html.find('class="group pdf-capture-page period-group"')
    if detail_idx < 0:
        detail_idx = html.find('class="group pdf-capture-page"')
    assert dash_idx != -1
    assert detail_idx != -1
    assert dash_idx < detail_idx
    assert "Payroll cost trend" in html or "Total payroll cost trajectory" in html or "Executive Summary" in html
    assert "<svg" in html
    assert "page-break-after: always" in html or "break-after: page" in html
    assert "Workforce Breakdown" in html or "Employment mix" in html or "category" in html.lower()
    assert "Avg Pay Rate" not in html
    assert "Avg Employer Cost" in html
    assert "Gross Payroll" in html or "Gross" in html
    assert "OT Premium" in html
    assert "OT Earnings" not in html
    assert ">Paid:" not in html and "Paid: $" not in html
    assert "Outstanding" not in html
    assert "Amount paid" not in html

    # PDF and Excel share identical Gross from analytics payload.
    maria = next(c for c in cats if c["worker_category"] == "contractor_1099")
    guiying = next(c for c in cats if c["worker_category"] == "temp")
    assert f"{maria['gross_pay']:.2f}" in html.replace(",", "")
    assert f"{guiying['gross_pay']:.2f}" in html.replace(",", "")

    xlsx = build_payroll_report_xlsx(report)
    assert xlsx[:2] == b"PK"
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(xlsx))
    assert "Workforce Breakdown" in wb.sheetnames
    wf = wb["Workforce Breakdown"]
    headers = [cell.value for cell in wf[2]]
    assert "Avg Pay Rate" not in headers
    assert "Avg Employer Cost / Hour" in headers
    assert "Regular/Base Earnings" in headers
    assert "OT Premium" in headers
    assert "OT Earnings" not in headers
    assert "Gross Payroll" in headers
    # Row values match analytics category payload (Base + OT Premium + Gross).
    by_label = {}
    for row in wf.iter_rows(min_row=3, values_only=True):
        if not row or not row[0] or row[0] == "Total":
            continue
        by_label[row[0]] = {
            "base_earnings": row[4],
            "ot_premium": row[5],
            "gross_pay": row[7],
        }
    assert by_label[maria["label"]]["base_earnings"] == round(maria["base_earnings"], 2)
    assert by_label[maria["label"]]["ot_premium"] == round(maria["ot_premium"], 2)
    assert by_label[maria["label"]]["gross_pay"] == round(maria["gross_pay"], 2)
    assert by_label[guiying["label"]]["base_earnings"] == round(guiying["base_earnings"], 2)
    assert by_label[guiying["label"]]["ot_premium"] == round(guiying["ot_premium"], 2)
    assert by_label[guiying["label"]]["gross_pay"] == round(guiying["gross_pay"], 2)
    pc = wb["Period Comparison"]
    pc_headers = [cell.value for cell in pc[2]]
    assert "Avg Pay Rate" not in pc_headers
    assert "Amount Paid" not in pc_headers
    assert "Outstanding" not in pc_headers
    assert "Avg Employer Cost / Hour" in pc_headers
    assert "Regular/Base Earnings" in pc_headers
    assert "OT Premium" in pc_headers
    assert "OT Earnings" not in pc_headers
    assert "% Δ Total Cost" in pc_headers
    assert "% Δ Hours" in pc_headers

    detail = wb["Payroll Reports"]
    detail_headers = [cell.value for cell in detail[5]]
    assert "Amount paid" not in detail_headers
    assert "Outstanding" not in detail_headers

    from backend.payroll_report import build_payroll_report_csv

    csv_bytes = build_payroll_report_csv(report)
    csv_text = csv_bytes.decode("utf-8-sig")
    assert "Amount paid" not in csv_text.splitlines()[0]
    assert "Outstanding" not in csv_text.splitlines()[0]
    assert "Gross pay" in csv_text.splitlines()[0]


def test_regular_and_ot_earnings_reconcile_to_gross():
    from backend.payroll_report_analytics import category_breakdown

    rows = [
        _row(
            name="W2",
            cat="w2",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=8.65,
            rate=17.0,
            ot_rate=25.5,
            line_id=1,
            user_id=1,
            employer_taxes=50,
            withheld=20,
        ),
        _row(
            name="Temp",
            cat="temp",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=12.49,
            rate=17.0,
            ot_rate=25.5,
            line_id=2,
            user_id=2,
        ),
        _row(
            name="1099",
            cat="contractor_1099",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=7.02,
            rate=17.0,
            ot_rate=25.5,
            line_id=3,
            user_id=3,
        ),
    ]
    rows[0]["total_payroll_cost"] = round(rows[0]["gross_pay"] + rows[0]["employer_taxes"], 2)
    cats = {c["worker_category"]: c for c in category_breakdown(rows)}
    for key in ("w2", "temp", "contractor_1099"):
        c = cats[key]
        assert abs(c["regular_earnings"] + c["ot_earnings"] + c["other_earnings"] - c["gross_pay"]) < 0.02
        assert c["mgmt_reconciles"] is True
        # Premium model still holds.
        assert abs(c["base_earnings"] + c["ot_premium"] + c["other_earnings"] - c["gross_pay"]) < 0.02
    assert cats["w2"]["avg_cost_per_hour"] > cats["w2"]["avg_pay_rate"]
    assert cats["temp"]["avg_pay_rate"] == cats["temp"]["avg_cost_per_hour"]


def test_employee_detail_redaction_flag():
    from backend.payroll_report_analytics import employee_summaries_by_category

    rows = [
        _row(
            name="Secret Worker",
            cat="w2",
            ps="2026-07-06",
            pe="2026-07-12",
            pay_date="2026-07-18",
            reg=40,
            ot=0,
            line_id=1,
            user_id=9,
        )
    ]
    hidden = employee_summaries_by_category(rows, include_identities=False)
    assert hidden == {}
    shown = employee_summaries_by_category(rows, include_identities=True)
    assert shown["w2"][0]["employee_name"] == "Secret Worker"

def test_ot_premium_included_in_gross_for_exports():
    row = _row(
        name="Maria Perez",
        cat="contractor_1099",
        ps="2026-06-22",
        pe="2026-06-28",
        pay_date="2026-07-04",
        reg=40,
        ot=4.65,
        line_id=1,
        user_id=1,
    )
    assert abs(row["base_earnings"] + row["ot_premium"] + row["other_earnings"] - row["gross_pay"]) < 0.02


def test_period_completeness_by_batch_terminal_status():
    """Complete = all batches paid/closed/finalized; category mix irrelevant."""
    from backend.payroll_report_analytics import (
        complete_period_keys_from_batches,
        period_batches_are_complete,
    )

    # W-2-only complete week
    w2_only = [
        {
            "pay_period_start": "2026-06-01",
            "pay_period_end": "2026-06-07",
            "worker_category": "w2",
            "status": "paid",
        }
    ]
    assert period_batches_are_complete(w2_only)
    assert complete_period_keys_from_batches(w2_only) == {("2026-06-01", "2026-06-07")}

    # Temp-only complete week (details finalized, not yet marked paid)
    temp_only = [
        {
            "pay_period_start": "2026-06-08",
            "pay_period_end": "2026-06-14",
            "worker_category": "temp",
            "status": "approved_for_payment",
            "payout_details_finalized_at": "2026-06-15T12:00:00",
        }
    ]
    assert period_batches_are_complete(temp_only)
    assert complete_period_keys_from_batches(temp_only) == {("2026-06-08", "2026-06-14")}

    # Mixed-category complete week
    mixed = [
        {
            "pay_period_start": "2026-06-15",
            "pay_period_end": "2026-06-21",
            "worker_category": "w2",
            "status": "paid",
        },
        {
            "pay_period_start": "2026-06-15",
            "pay_period_end": "2026-06-21",
            "worker_category": "temp",
            "status": "closed",
        },
        {
            "pay_period_start": "2026-06-15",
            "pay_period_end": "2026-06-21",
            "worker_category": "contractor_1099",
            "status": "paid",
        },
    ]
    assert period_batches_are_complete(mixed)
    assert complete_period_keys_from_batches(mixed) == {("2026-06-15", "2026-06-21")}

    # Partially finalized week — excluded
    partial = [
        {
            "pay_period_start": "2026-06-22",
            "pay_period_end": "2026-06-28",
            "worker_category": "w2",
            "status": "paid",
        },
        {
            "pay_period_start": "2026-06-22",
            "pay_period_end": "2026-06-28",
            "worker_category": "temp",
            "status": "draft",
        },
    ]
    assert not period_batches_are_complete(partial)
    assert complete_period_keys_from_batches(partial) == set()

    # Empty / open-only also incomplete
    assert not period_batches_are_complete([])
    assert not period_batches_are_complete(
        [
            {
                "pay_period_start": "2026-07-01",
                "pay_period_end": "2026-07-07",
                "worker_category": "w2",
                "status": "hours_reviewed",
            }
        ]
    )

    # Multi-period: only complete keys returned
    keys = complete_period_keys_from_batches(w2_only + temp_only + mixed + partial)
    assert keys == {
        ("2026-06-01", "2026-06-07"),
        ("2026-06-08", "2026-06-14"),
        ("2026-06-15", "2026-06-21"),
    }


def test_list_org_periods_complete_sql_filters_open_batches():
    """SQL path for terminal periods uses terminal/finalized HAVING clause."""
    from backend.payroll_report_analytics import list_org_periods_asc

    captured = {}

    class _Cur:
        def execute(self, sql, params=None):
            captured["sql"] = sql
            captured["params"] = params

        def fetchall(self):
            return [
                {"pay_period_start": "2026-06-01", "pay_period_end": "2026-06-07"},
            ]

    class _Conn:
        def cursor(self, dictionary=False):
            return _Cur()

    # Skip coverage post-filter so this unit test only asserts the SQL gate.
    periods = list_org_periods_asc(
        _Conn(), 3, require_complete=True, require_work_coverage=False
    )
    assert periods == [("2026-06-01", "2026-06-07")]
    sql = captured["sql"].lower()
    assert "payout_details_finalized_at is not null" in sql
    assert "'paid'" in sql and "'closed'" in sql
    assert "worker_category = 'w2'" not in sql
    assert captured["params"] == (3,)
