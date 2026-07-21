"""OT premium display helpers and payroll report presentation tests.

Does not change stored gross — only how OT is labeled for register/report/exports.
"""

from io import BytesIO
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from backend.payroll_overtime import (
    compute_earnings_breakdown,
    compute_overtime_premium,
    compute_wage_with_overtime,
    earnings_breakdown_from_line,
)
from backend.payroll_report import (
    DATE_MATCH_RULE,
    build_payroll_report_html,
    build_payroll_report_xlsx,
    build_report_row,
)


def test_ot_premium_time_and_a_half_example():
    """$20/hr × 1 OT hour @ $30 → premium $10 (not $30)."""
    premium = compute_overtime_premium(1, 20, 30)
    assert float(premium) == 10.0
    premium_half = compute_overtime_premium(1, 20)  # default 1.5×
    assert float(premium_half) == 10.0


def test_no_overtime_premium_is_zero():
    assert float(compute_overtime_premium(0, 20, 30)) == 0.0
    br = compute_earnings_breakdown(regular_hours=40, ot_hours=0, regular_rate=20, gross_pay=800)
    assert float(br["ot_premium"]) == 0.0
    assert float(br["base_earnings"]) == 800.0
    assert float(br["gross_pay"]) == 800.0
    assert abs(br["base_earnings"] + br["ot_premium"] + br["other_earnings"] - br["gross_pay"]) < 0.01


def test_ot_hours_premium_and_reconciliation():
    # 40 reg + 5 OT @ $20 / $30
    gross = float(compute_wage_with_overtime(40, 5, 20, 30))
    assert gross == 950.0  # 800 + 150
    br = compute_earnings_breakdown(
        regular_hours=40, ot_hours=5, regular_rate=20, ot_rate=30, gross_pay=gross
    )
    assert float(br["base_earnings"]) == 900.0  # 45 × 20
    assert float(br["ot_premium"]) == 50.0  # 5 × 10
    assert float(br["other_earnings"]) == 0.0
    assert float(br["gross_pay"]) == 950.0
    assert abs(br["base_earnings"] + br["ot_premium"] + br["other_earnings"] - br["gross_pay"]) < 0.01


def test_different_regular_and_ot_rates():
    br = compute_earnings_breakdown(
        regular_hours=40, ot_hours=2, regular_rate=17, ot_rate=25.50, gross_pay=731.0
    )
    assert float(br["ot_premium"]) == 17.0  # 2 × 8.50
    assert float(br["base_earnings"]) == 714.0  # 42 × 17
    assert abs(br["base_earnings"] + br["ot_premium"] + br["other_earnings"] - br["gross_pay"]) < 0.01


def test_1099_time_and_a_half_after_40():
    """1099 contractor receiving 1.5× after 40 — same premium math."""
    br = compute_earnings_breakdown(
        regular_hours=40, ot_hours=8.65, regular_rate=17, ot_rate=25.50, gross_pay=900.58
    )
    assert float(br["ot_premium"]) == 73.53  # 8.65 × 8.50
    assert float(br["base_earnings"]) == 827.05  # 48.65 × 17
    assert abs(br["base_earnings"] + br["ot_premium"] + br["other_earnings"] - 900.58) < 0.02
    # Gross calculation itself unchanged
    assert float(compute_wage_with_overtime(40, 8.65, 17, 25.50)) == 900.58


def test_earnings_breakdown_from_line_matches_gross():
    line = {
        "approved_hours": 40,
        "ot_hours": 5,
        "rate": 20,
        "ot_rate": 30,
        "gross_amount": 950,
        "sick_pay_amount": 0,
        "bonus_tip_amount": 0,
        "reimbursement_amount": 0,
        "adjustments": 0,
    }
    br = earnings_breakdown_from_line(line)
    assert float(br["ot_premium"]) == 50.0
    assert abs(br["base_earnings"] + br["ot_premium"] + br["other_earnings"] - br["gross_pay"]) < 0.01


def test_other_earnings_included_in_reconciliation():
    line = {
        "approved_hours": 40,
        "ot_hours": 0,
        "rate": 20,
        "gross_amount": 850,
        "sick_pay_amount": 50,
    }
    br = earnings_breakdown_from_line(line)
    assert float(br["base_earnings"]) == 800.0
    assert float(br["ot_premium"]) == 0.0
    assert float(br["other_earnings"]) == 50.0
    assert float(br["gross_pay"]) == 850.0


def test_report_row_ot_premium_and_labels():
    batch = {
        "id": 1,
        "batch_name": "W-2 week",
        "worker_category": "w2",
        "pay_period_start": "2026-07-06",
        "pay_period_end": "2026-07-12",
        "status": "approved_for_payment",
        "payout_details_finalized_at": "2026-07-13",
        "official_pay_date": "2026-07-15",
    }
    line = {
        "id": 10,
        "user_id": 99,
        "worker_name_snapshot": "Ada",
        "approved_hours": 40,
        "ot_hours": 1,
        "rate": 20,
        "ot_rate": 30,
        "gross_amount": 830,
        "payment_status": "paid",
        "payout_details_json": '{"payment":{"date":"2026-07-15"},"employee_deductions":{"fit":50},"employer_taxes":{"er_ss":51.46}}',
    }
    row = build_report_row(batch, line)
    assert row["ot_premium"] == 10.0
    assert row["base_earnings"] == 820.0
    assert abs(row["base_earnings"] + row["ot_premium"] + row["other_earnings"] - row["gross_pay"]) < 0.01
    assert row["pay_date"] == "2026-07-15"
    # Finalized date prefers Official Pay Date (business date), not system stamp.
    assert row["finalized_date"] == "2026-07-15"
    assert row["total_payroll_cost"] == round(830 + 51.46, 2)
    assert "W-2" in row["employee_category"]


def test_finalized_date_prefers_official_pay_date_over_system_stamp():
    """Historical finalize stamps NOW(); report must show the business pay date."""
    batch = {
        "id": 28,
        "batch_name": "TEMP-2026-004",
        "worker_category": "temp",
        "pay_period_start": "2026-05-25",
        "pay_period_end": "2026-05-31",
        "status": "paid",
        "payout_details_finalized_at": "2026-07-21T22:32:47",
        "official_pay_date": "2026-06-06",
    }
    line = {
        "id": 198,
        "user_id": 1,
        "worker_name_snapshot": "Aaliyah Rudowitz",
        "approved_hours": 10,
        "ot_hours": 0,
        "rate": 20,
        "ot_rate": 0,
        "gross_amount": 200,
        "payment_status": "pending",
        "payout_details_json": '{"payment":{"date":"2026-06-06"}}',
    }
    row = build_report_row(batch, line)
    assert row["pay_date"] == "2026-06-06"
    assert row["finalized_date"] == "2026-06-06"
    assert row["payroll_status"] == "Paid"
    # Batch is paid → report Payment status is Paid even if line is still pending.
    assert row["payment_status"] == "Paid"
    assert row["payment_status_key"] == "paid"


def test_payment_status_paid_when_batch_paid_but_line_approved_unpaid():
    batch = {
        "id": 18,
        "batch_name": "W2-2026-005",
        "worker_category": "w2",
        "pay_period_start": "2026-06-08",
        "pay_period_end": "2026-06-14",
        "status": "paid",
        "payout_details_finalized_at": "2026-06-21",
        "official_pay_date": "2026-06-13",
    }
    line = {
        "id": 50,
        "user_id": 2,
        "worker_name_snapshot": "Worker",
        "approved_hours": 40,
        "ot_hours": 0,
        "rate": 18,
        "ot_rate": 0,
        "gross_amount": 720,
        "payment_status": "approved_unpaid",
        "payout_details_json": None,
    }
    row = build_report_row(batch, line)
    assert row["payment_status"] == "Paid"
    assert row["finalized_date"] == "2026-06-13"

def test_excel_export_includes_totals_and_ot_premium():
    report = {
        "filters": {"all_history": True, "worker_category": "all", "report_type": "all_history"},
        "date_match_rule": DATE_MATCH_RULE,
        "report_type": "all_history",
        "rows": [
            {
                "employee_name": "Ada",
                "employee_category": "W-2 Employee",
                "batch_name": "W2",
                "batch_id": 1,
                "pay_period_start": "2026-07-06",
                "pay_period_end": "2026-07-12",
                "payroll_period": "2026-07-06 – 2026-07-12",
                "pay_date": "2026-07-15",
                "finalized_date": "2026-07-13",
                "regular_hours": 40,
                "ot_hours": 1,
                "base_earnings": 820,
                "ot_premium": 10,
                "other_earnings": 0,
                "gross_pay": 830,
                "employee_tax_deductions": 50,
                "other_deductions": 0,
                "net_pay": 780,
                "employer_taxes": 60,
                "total_payroll_cost": 890,
                "payment_status": "Paid",
                "payroll_status": "Ready To Pay",
            }
        ],
        "totals": {
            "regular_hours": 40,
            "ot_hours": 1,
            "base_earnings": 820,
            "ot_premium": 10,
            "other_earnings": 0,
            "gross_pay": 830,
            "employee_tax_deductions": 50,
            "other_deductions": 0,
            "net_pay": 780,
            "employer_taxes": 60,
            "total_payroll_cost": 890,
        },
        "summary": {
            "batch_count": 1,
            "unique_employees": 1,
            "gross_pay": 830,
            "employer_taxes": 60,
            "total_payroll_cost": 890,
        },
    }
    data = build_payroll_report_xlsx(report)
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    # Header row is after title/meta/summary lines
    headers = None
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and "OT premium" in row:
            headers = list(row)
            header_row_idx = i
            break
    assert headers is not None
    assert "OT premium" in headers
    assert "Regular/Base earnings" in headers
    assert "Total payroll cost" in headers
    found_total = False
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row and row[0] == "Totals":
            found_total = True
            ot_idx = headers.index("OT premium")
            assert float(row[ot_idx]) == 10.0
            gross_idx = headers.index("Gross pay")
            assert float(row[gross_idx]) == 830.0
            cost_idx = headers.index("Total payroll cost")
            assert float(row[cost_idx]) == 890.0
    assert found_total


def test_pdf_html_export_includes_totals_and_premium_note():
    report = {
        "filters": {
            "date_from": "2026-07-01",
            "date_to": "2026-07-31",
            "all_history": False,
            "worker_category": "all",
            "report_type": "custom_range",
            "date_basis": "pay_date",
        },
        "date_match_rule": DATE_MATCH_RULE,
        "report_type": "custom_range",
        "date_basis": "pay_date",
        "rows": [
            {
                "employee_name": "Bob",
                "employee_category": "1099 Contractor",
                "payroll_period": "2026-07-06 – 2026-07-12",
                "pay_period_start": "2026-07-06",
                "pay_period_end": "2026-07-12",
                "pay_date": "2026-07-14",
                "pay_date_display": "2026-07-14",
                "regular_hours": 40,
                "ot_hours": 2,
                "base_earnings": 840,
                "ot_premium": 20,
                "other_earnings": 0,
                "gross_pay": 860,
                "employee_tax_deductions": 0,
                "other_deductions": 0,
                "net_pay": 860,
                "employer_taxes": 0,
                "total_payroll_cost": 860,
                "payment_status": "Pending",
                "payroll_status": "Draft",
            }
        ],
        "totals": {
            "regular_hours": 40,
            "ot_hours": 2,
            "base_earnings": 840,
            "ot_premium": 20,
            "other_earnings": 0,
            "gross_pay": 860,
            "employee_tax_deductions": 0,
            "other_deductions": 0,
            "net_pay": 860,
            "employer_taxes": 0,
            "total_payroll_cost": 860,
        },
        "summary": {
            "batch_count": 1,
            "unique_employees": 1,
            "gross_pay": 860,
            "total_payroll_cost": 860,
        },
    }
    html = build_payroll_report_html(report)
    assert "OT Premium" in html or "OT premium" in html
    assert "$20.00" in html
    assert "Grand Total" in html
    assert "$860.00" in html
    assert "additional amount" in html.lower() or "regular hourly rate" in html.lower()


def test_register_gross_matches_payroll_record():
    """Register display gross equals stored line gross (no recalculation drift)."""
    line = {
        "approved_hours": 40,
        "ot_hours": 8.65,
        "rate": 17,
        "ot_rate": 25.50,
        "gross_amount": 900.58,
    }
    br = earnings_breakdown_from_line(line)
    assert float(br["gross_pay"]) == 900.58
    assert abs(br["base_earnings"] + br["ot_premium"] + br["other_earnings"] - 900.58) < 0.02


def test_date_range_rule_documented():
    """Default custom-range rule is Pay Date only (not combined OR)."""
    from backend.payroll_report import (
        DATE_MATCH_RULE_PERIOD_OVERLAP,
        date_match_rule_text,
    )

    assert "official pay date" in DATE_MATCH_RULE.lower()
    assert "overlap" not in DATE_MATCH_RULE.lower()
    assert " or " not in DATE_MATCH_RULE.lower()
    assert "overlap" in DATE_MATCH_RULE_PERIOD_OVERLAP.lower()
    assert "overlap" in date_match_rule_text("custom_range", date_basis="period_overlap").lower()
    assert "official pay date" in date_match_rule_text(
        "custom_range", date_basis="pay_date"
    ).lower()


def test_ot_premium_edge_cases_never_negative():
    # Missing OT rate → time-and-a-half premium
    br = compute_earnings_breakdown(regular_hours=40, ot_hours=2, regular_rate=20, ot_rate=None, gross_pay=860)
    assert float(br["ot_premium"]) == 20.0
    assert br["ot_premium"] >= 0

    # OT rate equal to regular → premium 0
    br_eq = compute_earnings_breakdown(
        regular_hours=40, ot_hours=2, regular_rate=20, ot_rate=20, gross_pay=840
    )
    assert float(br_eq["ot_premium"]) == 0.0
    assert abs(br_eq["base_earnings"] + br_eq["ot_premium"] + br_eq["other_earnings"] - 840) < 0.01

    # OT rate lower than regular → premium 0, base uses actual OT wages
    br_low = compute_earnings_breakdown(
        regular_hours=40, ot_hours=10, regular_rate=20, ot_rate=15, gross_pay=950
    )
    assert float(br_low["ot_premium"]) == 0.0
    assert float(br_low["base_earnings"]) == 950.0
    assert abs(br_low["base_earnings"] + br_low["ot_premium"] + br_low["other_earnings"] - 950) < 0.01

    # Zero / negative OT hours
    assert float(compute_overtime_premium(0, 20, 30)) == 0.0
    assert float(compute_overtime_premium(-5, 20, 30)) == 0.0

    # Salaried / non-hourly
    br_sal = compute_earnings_breakdown(
        regular_hours=0, ot_hours=0, regular_rate=0, gross_pay=1200
    )
    assert float(br_sal["ot_premium"]) == 0.0
    assert float(br_sal["base_earnings"]) == 0.0
    assert float(br_sal["other_earnings"]) == 1200.0


def test_excel_uses_numeric_and_date_cells():
    from datetime import date

    report = {
        "filters": {
            "date_from": "2026-07-01",
            "date_to": "2026-07-31",
            "report_type": "custom_range",
            "date_basis": "pay_date",
        },
        "date_match_rule": DATE_MATCH_RULE,
        "report_type": "custom_range",
        "rows": [
            {
                "employee_name": "Ada",
                "employee_category": "W-2 Employee",
                "batch_name": "W2",
                "batch_id": 1,
                "pay_period_start": "2026-07-06",
                "pay_period_end": "2026-07-12",
                "payroll_period": "2026-07-06 – 2026-07-12",
                "pay_date": "2026-07-15",
                "finalized_date": "2026-07-13",
                "regular_hours": 40,
                "ot_hours": 1,
                "base_earnings": 820.0,
                "ot_premium": 10.0,
                "other_earnings": 0.0,
                "gross_pay": 830.0,
                "employee_tax_deductions": 50.5,
                "other_deductions": 0.0,
                "net_pay": 779.5,
                "employer_taxes": 60.0,
                "total_payroll_cost": 890.0,
                "payment_status": "Paid",
                "payroll_status": "Ready To Pay",
            }
        ],
        "totals": {
            "regular_hours": 40,
            "ot_hours": 1,
            "base_earnings": 820.0,
            "ot_premium": 10.0,
            "other_earnings": 0.0,
            "gross_pay": 830.0,
            "employee_tax_deductions": 50.5,
            "other_deductions": 0.0,
            "net_pay": 779.5,
            "employer_taxes": 60.0,
            "total_payroll_cost": 890.0,
        },
        "summary": {"batch_count": 1, "unique_employees": 1, "total_payroll_cost": 890.0},
    }
    data = build_payroll_report_xlsx(report)
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    headers = None
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and "Pay Date" in row:
            headers = list(row)
            header_row_idx = i
            break
    assert headers is not None
    pay_date_idx = headers.index("Pay Date") + 1
    gross_idx = headers.index("Gross pay") + 1
    ot_idx = headers.index("OT premium") + 1
    data_row = header_row_idx + 1
    cell_date = ws.cell(row=data_row, column=pay_date_idx)
    cell_gross = ws.cell(row=data_row, column=gross_idx)
    cell_ot = ws.cell(row=data_row, column=ot_idx)
    assert isinstance(cell_date.value, date)
    assert isinstance(cell_gross.value, (int, float))
    assert not isinstance(cell_gross.value, str)
    assert float(cell_ot.value) == 10.0
    assert float(ws.cell(row=data_row + 1, column=gross_idx).value) == 830.0


def test_pdf_and_screen_totals_match_exactly():
    totals = {
        "regular_hours": 80.5,
        "ot_hours": 3.25,
        "base_earnings": 1670.0,
        "ot_premium": 32.5,
        "other_earnings": 10.0,
        "gross_pay": 1712.5,
        "employee_tax_deductions": 100.0,
        "other_deductions": 5.0,
        "net_pay": 1607.5,
        "employer_taxes": 80.0,
        "total_payroll_cost": 1792.5,
    }
    report = {
        "filters": {
            "date_from": "2026-07-01",
            "date_to": "2026-07-31",
            "report_type": "custom_range",
            "date_basis": "pay_date",
        },
        "date_match_rule": DATE_MATCH_RULE,
        "report_type": "custom_range",
        "rows": [],
        "totals": totals,
        "summary": {"batch_count": 0, "unique_employees": 0, **totals},
    }
    html = build_payroll_report_html(report)
    assert "thead" in html and "table-header-group" in html
    assert "page-break-inside: avoid" in html
    assert "Grand Total" in html
    assert "$1,712.50" in html
    assert "$1,792.50" in html
    xlsx = build_payroll_report_xlsx({**report, "rows": []})
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    headers = None
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and "Gross pay" in row:
            headers = list(row)
            header_row_idx = i
            break
    assert headers is not None
    gross_idx = headers.index("Gross pay") + 1
    assert float(ws.cell(row=header_row_idx + 1, column=gross_idx).value) == 1712.5


def test_date_range_match_does_not_duplicate_line():
    """Same line_id twice from SQL → still one row; pay_date basis uses official_pay_date."""
    from backend.payroll_report import query_payroll_report

    conn = MagicMock()
    cursor = MagicMock()
    conn.cursor.return_value = cursor
    sample = {
        "batch_id": 1,
        "batch_name": "W2",
        "worker_category": "w2",
        "pay_period_start": "2026-07-06",
        "pay_period_end": "2026-07-12",
        "batch_status": "paid",
        "payout_details_finalized_at": "2026-07-13",
        "official_pay_date": "2026-07-10",
        "line_id": 99,
        "user_id": 5,
        "worker_name_snapshot": "Eve",
        "approved_hours": 40,
        "ot_hours": 0,
        "rate": 20,
        "ot_rate": 0,
        "gross_amount": 800,
        "total_amount": 800,
        "gross_wages": 800,
        "sick_pay_amount": 0,
        "bonus_tip_amount": 0,
        "reimbursement_amount": 0,
        "adjustments": 0,
        "payment_status": "paid",
        "net_pay": 800,
        "payout_details_json": '{"payment":{"date":"2026-07-10"}}',
    }
    cursor.fetchall.return_value = [sample, dict(sample)]
    with patch("backend.payroll_operations.ensure_payout_batches_tables"), patch(
        "backend.payroll_report.table_has_column", return_value=True
    ):
        report = query_payroll_report(
            conn,
            3,
            date_from="2026-07-06",
            date_to="2026-07-12",
            date_basis="pay_date",
        )
    assert report["count"] == 1
    assert len(report["rows"]) == 1
    assert report["report_type"] == "custom_range"
    assert "overlap" not in (report["date_match_rule"] or "").lower()


def test_query_payroll_report_filters_periods_and_categories():
    """Smoke-test query assembly with a mocked cursor."""
    from backend.payroll_report import query_payroll_report

    conn = MagicMock()
    cursor = MagicMock()
    conn.cursor.return_value = cursor
    cursor.fetchall.return_value = [
        {
            "batch_id": 1,
            "batch_name": "1099",
            "worker_category": "contractor_1099",
            "pay_period_start": "2026-07-06",
            "pay_period_end": "2026-07-12",
            "batch_status": "hours_reviewed",
            "payout_details_finalized_at": None,
            "official_pay_date": None,
            "line_id": 2,
            "user_id": 5,
            "worker_name_snapshot": "Cara",
            "approved_hours": 40,
            "ot_hours": 4,
            "rate": 20,
            "ot_rate": 30,
            "gross_amount": 920,
            "total_amount": 920,
            "gross_wages": 920,
            "sick_pay_amount": 0,
            "bonus_tip_amount": 0,
            "reimbursement_amount": 0,
            "adjustments": 0,
            "payment_status": "pending",
            "net_pay": 920,
            "payout_details_json": None,
        },
        {
            "batch_id": 2,
            "batch_name": "W2 other week",
            "worker_category": "w2",
            "pay_period_start": "2026-06-29",
            "pay_period_end": "2026-07-05",
            "batch_status": "paid",
            "payout_details_finalized_at": "2026-07-06",
            "official_pay_date": "2026-07-08",
            "line_id": 3,
            "user_id": 6,
            "worker_name_snapshot": "Dan",
            "approved_hours": 40,
            "ot_hours": 0,
            "rate": 18,
            "ot_rate": 0,
            "gross_amount": 720,
            "total_amount": 720,
            "gross_wages": 720,
            "sick_pay_amount": 0,
            "bonus_tip_amount": 0,
            "reimbursement_amount": 0,
            "adjustments": 0,
            "payment_status": "paid",
            "net_pay": 700,
            "payout_details_json": '{"payment":{"date":"2026-07-08"}}',
        },
    ]
    with patch("backend.payroll_operations.ensure_payout_batches_tables"), patch(
        "backend.payroll_report.table_has_column", return_value=True
    ):
        multi = query_payroll_report(
            conn,
            3,
            period_starts=["2026-07-06"],
            period_ends=["2026-07-12"],
        )
        assert "rows" in multi
        assert multi["report_type"] == "payroll_period"
        assert multi["date_match_rule"]

        # Default date_basis=pay_date: only Dan (official_pay_date in range)
        ranged = query_payroll_report(
            conn,
            3,
            date_from="2026-07-07",
            date_to="2026-07-10",
            date_basis="pay_date",
        )
        names = {r["employee_name"] for r in ranged["rows"]}
        assert names == {"Dan"}
        assert "Cara" not in names

        # period_overlap: Cara's period overlaps Jul 7–10
        overlapped = query_payroll_report(
            conn,
            3,
            date_from="2026-07-07",
            date_to="2026-07-10",
            date_basis="period_overlap",
        )
        overlap_names = {r["employee_name"] for r in overlapped["rows"]}
        assert "Cara" in overlap_names
        for r in overlapped["rows"]:
            assert abs(
                r["base_earnings"] + r["ot_premium"] + r["other_earnings"] - r["gross_pay"]
            ) < 0.02
