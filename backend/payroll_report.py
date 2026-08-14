"""Cross-period Payroll Report: query, Excel, and PDF/HTML exports.

OT amounts use premium-only presentation (see payroll_overtime.earnings_breakdown_from_line).
Does not mutate stored gross or historical payroll amounts.

Phase 1: batch official_pay_date is the sole reporting Pay Date (no period-end fallback).
"""

from __future__ import annotations

import calendar
import io
import json
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Optional

from backend.payroll_operations import BATCH_STATUSES, CATEGORY_LABELS, WORKER_CATEGORIES
from backend.payroll_overtime import earnings_breakdown_from_line
from backend.payroll_status_display import (
    DISPLAY_STATUS_LABELS,
    compute_display_status,
)
from backend.ta_helpers import json_safe, table_has_column


# Default custom-range rule (Pay Date basis). Never the legacy combined OR.
DATE_MATCH_RULE = (
    "Includes rows where the official Pay Date falls within the selected range."
)
DATE_MATCH_RULE_PERIOD_OVERLAP = (
    "Includes rows where the pay period overlaps the selected range."
)
DATE_MATCH_RULE_MONTHLY = (
    "Includes rows whose official Pay Date falls in the selected month and year. "
    "Rows without an official Pay Date are excluded."
)
DATE_MATCH_RULE_PAYROLL_PERIOD = "Includes rows for the selected payroll period(s)."
DATE_MATCH_RULE_ALL_HISTORY = (
    "Includes all payroll history. Rows without an official Pay Date show Pay Date Missing."
)

REPORT_TYPES = (
    "payroll_period",
    "monthly_paid",
    "custom_range",
    "all_history",
)

DATE_BASES = ("pay_date", "period_overlap")

# Record-level columns (Excel / JSON). PDF uses a readable subset.
REPORT_COLUMNS = (
    ("employee_name", "Employee"),
    ("employee_category", "Category"),
    ("batch_name", "Payroll batch"),
    ("batch_id", "Batch ID"),
    ("pay_period_start", "Payroll-period start"),
    ("pay_period_end", "Payroll-period end"),
    ("pay_date", "Pay Date"),
    ("finalized_date", "Finalized date"),
    ("regular_hours", "Regular hours"),
    ("ot_hours", "OT hours"),
    ("base_earnings", "Regular/Base earnings"),
    ("ot_premium", "OT premium"),
    ("other_earnings", "Other earnings"),
    ("gross_pay", "Gross pay"),
    ("employee_tax_deductions", "Employee taxes"),
    ("other_deductions", "Deductions"),
    ("net_pay", "Net pay"),
    ("employer_taxes", "Employer taxes"),
    ("total_payroll_cost", "Total payroll cost"),
    ("payroll_status", "Payroll status"),
    ("payment_status", "Payment status"),
)

PDF_COLUMNS = (
    ("employee_name", "Employee"),
    ("employee_category", "Category"),
    ("pay_date_display", "Pay Date"),
    ("regular_hours", "Reg hrs"),
    ("ot_hours", "OT hrs"),
    ("base_earnings", "Base"),
    ("ot_premium", "OT prem"),
    ("gross_pay", "Gross"),
    ("employee_tax_deductions", "EE taxes"),
    ("net_pay", "Net"),
    ("employer_taxes", "ER taxes"),
    ("total_payroll_cost", "Total cost"),
    ("payment_status", "Payment"),
)

HOUR_TOTAL_KEYS = ("regular_hours", "ot_hours")
MONEY_TOTAL_KEYS = (
    "base_earnings",
    "ot_premium",
    "regular_earnings",
    "ot_earnings",
    "other_earnings",
    "gross_pay",
    "employee_tax_deductions",
    "other_deductions",
    "net_pay",
    "amount_paid",
    "outstanding_balance",
    "employer_taxes",
    "total_payroll_cost",
)

DATE_CELL_KEYS = (
    "pay_date",
    "pay_period_start",
    "pay_period_end",
    "finalized_date",
    "official_pay_date",
)


def date_match_rule_text(
    report_type: str,
    *,
    date_basis: str = "pay_date",
) -> str:
    rt = str(report_type or "").strip().lower()
    if rt == "monthly_paid":
        return DATE_MATCH_RULE_MONTHLY
    if rt == "payroll_period":
        return DATE_MATCH_RULE_PAYROLL_PERIOD
    if rt == "all_history":
        return DATE_MATCH_RULE_ALL_HISTORY
    if rt == "custom_range":
        basis = str(date_basis or "pay_date").strip().lower()
        if basis == "period_overlap":
            return DATE_MATCH_RULE_PERIOD_OVERLAP
        return DATE_MATCH_RULE
    return DATE_MATCH_RULE


def _parse_ymd(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()[:10]
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _money(val: Any) -> float:
    try:
        return float(Decimal(str(val if val is not None else 0)))
    except Exception:
        return 0.0


def _parse_details(raw: Any) -> dict:
    if isinstance(raw, dict):
        return raw
    if not raw:
        return {}
    try:
        parsed = json.loads(raw) if isinstance(raw, str) else raw
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _batch_official_pay_date(batch: dict) -> Optional[str]:
    """Reporting Pay Date from batch official_pay_date only — no period-end fallback."""
    d = _parse_ymd(batch.get("official_pay_date"))
    return d.isoformat() if d else None


def _periods_overlap(ps: Optional[date], pe: Optional[date], start: date, end: date) -> bool:
    if not ps and not pe:
        return False
    period_start = ps or pe
    period_end = pe or ps
    assert period_start is not None and period_end is not None
    return period_start <= end and period_end >= start


def _row_matches_pay_date(row: dict, start: date, end: date) -> bool:
    pd = _parse_ymd(row.get("pay_date") or row.get("official_pay_date"))
    if not pd:
        return False
    return start <= pd <= end


def _row_matches_period_overlap(row: dict, start: date, end: date) -> bool:
    ps = _parse_ymd(row.get("pay_period_start"))
    pe = _parse_ymd(row.get("pay_period_end"))
    return _periods_overlap(ps, pe, start, end)


def _row_matches_date_range(
    row: dict, start: date, end: date, *, date_basis: str = "pay_date"
) -> bool:
    """Custom range match by explicit basis only (never combined OR)."""
    basis = str(date_basis or "pay_date").strip().lower()
    if basis == "period_overlap":
        return _row_matches_period_overlap(row, start, end)
    return _row_matches_pay_date(row, start, end)


def _employee_tax_total(details: dict) -> float:
    """Employee taxes shown on the report.

    When settlement records what was actually withheld/paid, use amount_withheld
    so Net (amount paid) reconciles: Gross − EE taxes ≈ Net.

    paid_full_gross_without_withholding pays the full gross and withholds $0 —
    estimated employee_deductions remain a tax *balance owed*, not taxes taken
    from this paycheck, so the report must show $0 EE taxes (not the estimate).
    Otherwise Gross=Net while EE taxes > 0 looks like a broken Net column.
    """
    settlement = details.get("settlement") or {}
    if bool(settlement.get("paid_full_gross_without_withholding")):
        return 0.0
    amount_paid = settlement.get("amount_paid")
    has_paid = amount_paid is not None and str(amount_paid).strip() != ""
    if has_paid or settlement.get("amount_withheld") is not None:
        if has_paid or _money(settlement.get("amount_withheld")) > 0:
            return round(_money(settlement.get("amount_withheld")), 2)
    ded = details.get("employee_deductions") or {}
    return round(sum(_money(v) for v in ded.values()), 2)


def _report_net_pay(
    line: dict,
    details: dict,
    *,
    gross_pay: float,
    emp_tax: float,
    other_ded: float,
    worker_category: Optional[str] = None,
) -> float:
    """Net for the report.

    Temp/1099: Net = earned Gross − EE taxes − other deductions. Amount paid is a
    separate column (may be less than Gross when OT catch-up is outstanding).

    W-2: prefer settlement amount_paid when present (cash net after withholding).
    """
    cat = str(worker_category or line.get("worker_category") or "").strip()
    if cat in ("temp", "contractor_1099", "tryout"):
        return round(_money(gross_pay) - emp_tax - other_ded, 2)

    settlement = details.get("settlement") or {}
    amount_paid = settlement.get("amount_paid")
    if amount_paid is not None and str(amount_paid).strip() != "":
        return round(_money(amount_paid), 2)
    if bool(settlement.get("paid_full_gross_without_withholding")) and gross_pay > 0:
        return round(_money(gross_pay), 2)
    net = line.get("net_pay")
    if net is not None and str(net).strip() != "":
        return round(_money(net), 2)
    return round(_money(gross_pay) - emp_tax - other_ded, 2)


def _employer_tax_total(details: dict) -> float:
    er = details.get("employer_taxes") or {}
    return round(sum(_money(v) for v in er.values()), 2)


def _other_deductions_total(details: dict) -> float:
    """Non-tax employee deductions if present (catch-up / misc)."""
    settlement = details.get("settlement") or {}
    catch_up = _money(settlement.get("catch_up_withholding"))
    misc = details.get("other_deductions") or details.get("voluntary_deductions") or {}
    misc_total = sum(_money(v) for v in misc.values()) if isinstance(misc, dict) else _money(misc)
    return round(catch_up + misc_total, 2)


def _payment_status_label(st: str) -> str:
    key = str(st or "pending").strip().lower()
    labels = {
        "paid": "Paid",
        "unpaid": "UNPAID",
        "partial": "Partial — balance due",
        "approved_unpaid": "Approved — unpaid",
        "pending": "Pending",
    }
    return labels.get(key, key.replace("_", " ").title() or "Pending")


def _effective_payment_status(batch: dict, line: dict, *, outstanding: float = 0.0) -> str:
    """Line payment status for reports.

    Explicit Unpaid always wins so finalized-unpaid records are never treated as
    paid, even when the batch workflow status is paid. Legacy rows without
    payment_recorded keep the prior rule: a paid batch's lines are Paid unless
    an outstanding OT/catch-up balance remains.
    """
    from backend.payroll_worker_categories import is_payment_recorded_unpaid, line_payment_recorded

    details = line.get("payout_details") or _parse_details(line.get("payout_details_json"))
    if is_payment_recorded_unpaid(line, details, batch):
        return "unpaid"
    recorded = line_payment_recorded(line, details, batch)
    if recorded == "paid":
        if float(outstanding or 0) > 0.005:
            return "partial"
        return "paid"
    if float(outstanding or 0) > 0.005:
        return "partial"
    batch_st = str(batch.get("status") or "").strip().lower()
    if batch_st == "paid":
        return "paid"
    return str(line.get("payment_status") or "pending").strip().lower() or "pending"


def _finalized_date_str(batch: dict) -> str:
    """Business Finalized date for reports.

    Prefer Official Pay Date (the date entered at finalization) over the
    system timestamp of the finalize click. Historical finalization stamps
    payout_details_finalized_at as "today", which mislabels older payroll
    (e.g. TEMP-2026-004 paid 2026-06-06 but finalized in-system on 2026-07-21).
    """
    pay_date = _batch_official_pay_date(batch)
    if pay_date:
        return pay_date
    raw = batch.get("payout_details_finalized_at")
    if not raw:
        return ""
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    s = str(raw).strip()
    if not s:
        return ""
    d = _parse_ymd(s[:10])
    return d.isoformat() if d else s[:19]


def build_report_row(batch: dict, line: dict, *, report_type: Optional[str] = None) -> dict[str, Any]:
    details = line.get("payout_details") or _parse_details(line.get("payout_details_json"))
    breakdown = earnings_breakdown_from_line(line)
    cat = str(batch.get("worker_category") or line.get("worker_category") or "")
    display_status = compute_display_status(batch)
    pay_date = _batch_official_pay_date(batch)
    pay_date_missing = pay_date is None
    emp_tax = _employee_tax_total(details)
    other_ded = _other_deductions_total(details)
    employer_taxes = _employer_tax_total(details)
    # Earned gross always includes OT premium (Base + OT prem [+ other] = Gross).
    gross_pay = round(float(breakdown["gross_pay"]), 2)
    settlement = details.get("settlement") or {}
    amount_paid_raw = settlement.get("amount_paid")
    outstanding = round(_money(settlement.get("outstanding_balance")), 2)
    if amount_paid_raw is not None and str(amount_paid_raw).strip() != "":
        amount_paid = round(_money(amount_paid_raw), 2)
    elif bool(settlement.get("paid_full_gross_without_withholding")):
        amount_paid = gross_pay
    else:
        amount_paid = None
    # Temp/1099 with zero withholding: EE taxes stay $0 for report when paid-full
    # or when amount_withheld is explicitly 0 / absent with no tax deductions taken.
    if cat in ("temp", "contractor_1099", "tryout"):
        emp_tax = round(_money(settlement.get("amount_withheld") or 0), 2)

    net = _report_net_pay(
        line,
        details,
        gross_pay=gross_pay,
        emp_tax=emp_tax,
        other_ded=other_ded,
        worker_category=cat,
    )

    ps = batch.get("pay_period_start")
    pe = batch.get("pay_period_end")
    period_label = ""
    if ps and pe:
        period_label = f"{str(ps)[:10]} – {str(pe)[:10]}"
    elif ps or pe:
        period_label = str(ps or pe)[:10]

    total_payroll_cost = round(_money(gross_pay) + employer_taxes, 2)

    payment_st = _effective_payment_status(batch, line, outstanding=outstanding)
    paid_for_totals = amount_paid if amount_paid is not None else 0.0
    if payment_st == "unpaid":
        paid_for_totals = 0.0
        outstanding = round(_money(amount_paid if amount_paid is not None else gross_pay), 2)
    vendor = details.get("vendor") if isinstance(details.get("vendor"), dict) else {}
    from backend.payroll_worker_categories import payment_vendor_display_name

    return {
        "line_id": line.get("id"),
        "batch_id": batch.get("id"),
        "user_id": line.get("user_id"),
        "employee_name": line.get("worker_name_snapshot") or "",
        "employee_category": CATEGORY_LABELS.get(cat, cat),
        "worker_category": cat,
        "vendor_name": payment_vendor_display_name(vendor.get("name")) if vendor.get("name") else "",
        "payroll_period": period_label,
        "pay_period_start": str(ps)[:10] if ps else "",
        "pay_period_end": str(pe)[:10] if pe else "",
        "pay_date": pay_date or "",
        "official_pay_date": pay_date or "",
        "pay_date_missing": pay_date_missing,
        "pay_date_display": pay_date if pay_date else "Pay Date Missing",
        "finalized_date": _finalized_date_str(batch),
        "regular_hours": breakdown["regular_hours"],
        "ot_hours": breakdown["ot_hours"],
        "regular_rate": breakdown["regular_rate"],
        "ot_rate": breakdown["ot_rate"],
        "base_earnings": breakdown["base_earnings"],
        "ot_premium": breakdown["ot_premium"],
        "regular_earnings": breakdown.get("regular_earnings", 0.0),
        "ot_earnings": breakdown.get("ot_earnings", 0.0),
        "other_earnings": breakdown["other_earnings"],
        "gross_pay": gross_pay,
        "earned_gross": gross_pay,
        "employee_tax_deductions": emp_tax,
        "other_deductions": other_ded,
        "net_pay": net,
        "amount_paid": paid_for_totals,
        "outstanding_balance": outstanding,
        "employer_taxes": employer_taxes,
        "total_payroll_cost": total_payroll_cost,
        "payment_status": _payment_status_label(payment_st),
        "payment_status_key": payment_st,
        "payroll_status": DISPLAY_STATUS_LABELS.get(display_status, display_status),
        "payroll_status_key": display_status,
        "batch_status": batch.get("status"),
        "batch_name": batch.get("batch_name") or "",
    }


def _sum_totals(rows: list[dict]) -> dict[str, float]:
    totals = {k: 0.0 for k in (*HOUR_TOTAL_KEYS, *MONEY_TOTAL_KEYS)}
    for row in rows:
        for k in totals:
            totals[k] = round(totals[k] + _money(row.get(k)), 2)
    return totals


def _build_summary(rows: list[dict], totals: dict[str, float]) -> dict[str, Any]:
    batch_ids = {r.get("batch_id") for r in rows if r.get("batch_id") is not None}
    user_ids = {r.get("user_id") for r in rows if r.get("user_id") is not None}
    pay_dates = {
        str(r.get("pay_date") or r.get("official_pay_date") or "")[:10]
        for r in rows
        if str(r.get("pay_date") or r.get("official_pay_date") or "").strip()
    }
    periods = {
        (
            str(r.get("pay_period_start") or "")[:10],
            str(r.get("pay_period_end") or "")[:10],
        )
        for r in rows
        if str(r.get("pay_period_start") or "").strip()
        and str(r.get("pay_period_end") or "").strip()
    }
    return {
        "batch_count": len(batch_ids),
        "unique_employees": len(user_ids),
        "payroll_period_count": len(periods),
        "official_pay_date_count": len(pay_dates),
        "total_hours": round(
            float(totals.get("regular_hours") or 0) + float(totals.get("ot_hours") or 0),
            2,
        ),
        **totals,
    }


def _infer_report_type(
    *,
    report_type: Optional[str],
    all_history: bool,
    month: Optional[int],
    year: Optional[int],
    period_pairs: list[tuple[str, str]],
    date_from: Optional[str],
    date_to: Optional[str],
) -> str:
    if report_type:
        rt = str(report_type).strip().lower()
        if rt in REPORT_TYPES:
            return rt
        raise ValueError(f"Invalid report_type: {report_type}")
    if all_history:
        return "all_history"
    if month is not None and year is not None:
        return "monthly_paid"
    if period_pairs:
        return "payroll_period"
    if date_from and date_to:
        return "custom_range"
    return "all_history"


def _normalize_date_basis(date_basis: Optional[str]) -> str:
    basis = str(date_basis or "pay_date").strip().lower()
    if basis not in DATE_BASES:
        raise ValueError("date_basis must be 'pay_date' or 'period_overlap'")
    return basis


def query_payroll_report(
    conn,
    organization_id: int,
    *,
    period_starts: Optional[list[str]] = None,
    period_ends: Optional[list[str]] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    all_history: bool = False,
    user_id: Optional[int] = None,
    worker_category: Optional[str] = None,
    payroll_status: Optional[str] = None,
    payment_status: Optional[str] = None,
    limit: int = 5000,
    report_type: Optional[str] = None,
    date_basis: str = "pay_date",
    month: Optional[int] = None,
    year: Optional[int] = None,
    comparison_range: Optional[int] = None,
    compare_with: Optional[str] = None,
    trend_range: Optional[int] = None,
    include_analytics: bool = True,
    include_employee_detail: bool = True,
) -> dict[str, Any]:
    """Return filtered payroll report rows + totals across categories/periods."""
    from backend.payroll_operations import ensure_payout_batches_tables

    ensure_payout_batches_tables(conn.cursor())
    c = conn.cursor(dictionary=True)
    has_details = table_has_column(c, "payout_batch_lines", "payout_details_json")
    has_official_pay_date = table_has_column(c, "payout_batches", "official_pay_date")

    select_cols = """
        pb.id AS batch_id,
        pb.batch_name,
        pb.worker_category,
        pb.pay_period_start,
        pb.pay_period_end,
        pb.status AS batch_status,
        pb.payout_details_finalized_at,
        pbl.id AS line_id,
        pbl.user_id,
        pbl.worker_name_snapshot,
        pbl.approved_hours,
        pbl.ot_hours,
        pbl.rate,
        pbl.ot_rate,
        pbl.gross_amount,
        pbl.total_amount,
        pbl.gross_wages,
        pbl.sick_pay_amount,
        pbl.bonus_tip_amount,
        pbl.reimbursement_amount,
        pbl.adjustments,
        pbl.payment_status,
        pbl.net_pay
    """
    if has_official_pay_date:
        select_cols += ", pb.official_pay_date"
    if has_details:
        select_cols += ", pbl.payout_details_json"

    period_pairs: list[tuple[str, str]] = []
    starts = [str(s).strip()[:10] for s in (period_starts or []) if str(s).strip()]
    ends = [str(e).strip()[:10] for e in (period_ends or []) if str(e).strip()]
    if starts and ends and len(starts) == len(ends):
        period_pairs = list(zip(starts, ends))
    elif starts and ends and len(starts) == 1 and len(ends) == 1:
        period_pairs = [(starts[0], ends[0])]

    month_i = int(month) if month is not None and str(month).strip() != "" else None
    year_i = int(year) if year is not None and str(year).strip() != "" else None

    resolved_type = _infer_report_type(
        report_type=report_type,
        all_history=all_history,
        month=month_i,
        year=year_i,
        period_pairs=period_pairs,
        date_from=date_from,
        date_to=date_to,
    )
    basis = _normalize_date_basis(date_basis)
    match_rule = date_match_rule_text(resolved_type, date_basis=basis)

    if resolved_type == "monthly_paid":
        if month_i is None or year_i is None:
            raise ValueError("monthly_paid requires month and year")
        if month_i < 1 or month_i > 12:
            raise ValueError("month must be 1–12")
        if not has_official_pay_date:
            # Column missing: nothing can match official pay date filters.
            empty_totals = _sum_totals([])
            return json_safe(
                {
                    "rows": [],
                    "totals": empty_totals,
                    "summary": _build_summary([], empty_totals),
                    "count": 0,
                    "report_type": resolved_type,
                    "date_basis": basis,
                    "excluded_missing_pay_date_count": 0,
                    "filters": {
                        "report_type": resolved_type,
                        "date_basis": basis,
                        "period_starts": starts,
                        "period_ends": ends,
                        "date_from": date_from or "",
                        "date_to": date_to or "",
                        "month": month_i,
                        "year": year_i,
                        "all_history": False,
                        "user_id": int(user_id) if user_id else None,
                        "worker_category": worker_category or "all",
                        "payroll_status": payroll_status or "all",
                        "payment_status": payment_status or "all",
                        "date_match_rule": match_rule,
                    },
                    "date_match_rule": match_rule,
                    "columns": [{"key": k, "label": lab} for k, lab in REPORT_COLUMNS],
                    "categories": [
                        {"value": "all", "label": "All categories"},
                        *[{"value": c, "label": CATEGORY_LABELS[c]} for c in WORKER_CATEGORIES],
                    ],
                    "payroll_statuses": [
                        {"value": "all", "label": "All payroll statuses"},
                        *[
                            {"value": k, "label": v}
                            for k, v in DISPLAY_STATUS_LABELS.items()
                        ],
                    ],
                    "payment_statuses": [
                        {"value": "all", "label": "All payment statuses"},
                        {"value": "pending", "label": "Pending"},
                        {"value": "approved_unpaid", "label": "Approved — unpaid"},
                        {"value": "paid", "label": "Paid"},
                    ],
                    "batch_statuses": list(BATCH_STATUSES),
                }
            )

    q = f"""
        SELECT {select_cols}
        FROM payout_batch_lines pbl
        INNER JOIN payout_batches pb ON pb.id = pbl.batch_id
        WHERE pb.organization_id = %s AND pbl.organization_id = %s
    """
    params: list[Any] = [int(organization_id), int(organization_id)]

    if resolved_type == "payroll_period" and period_pairs:
        placeholders = " OR ".join(
            ["(pb.pay_period_start = %s AND pb.pay_period_end = %s)"] * len(period_pairs)
        )
        q += f" AND ({placeholders})"
        for ps, pe in period_pairs:
            params.extend([ps, pe])
    elif resolved_type == "monthly_paid":
        q += (
            " AND pb.official_pay_date IS NOT NULL"
            " AND MONTH(pb.official_pay_date) = %s"
            " AND YEAR(pb.official_pay_date) = %s"
        )
        params.extend([month_i, year_i])

    if worker_category and worker_category != "all":
        if worker_category not in WORKER_CATEGORIES:
            raise ValueError("Invalid worker_category")
        q += " AND pb.worker_category = %s"
        params.append(worker_category)

    if user_id:
        q += " AND pbl.user_id = %s"
        params.append(int(user_id))

    if payment_status and payment_status != "all":
        q += " AND pbl.payment_status = %s"
        params.append(str(payment_status))

    if resolved_type == "monthly_paid" and has_official_pay_date:
        # Period primary, then pay date, then employee (matches analytics groups).
        q += (
            " ORDER BY pb.pay_period_start, pb.pay_period_end, pb.official_pay_date,"
            " pb.worker_category, pbl.worker_name_snapshot LIMIT %s"
        )
    else:
        q += " ORDER BY pb.pay_period_start DESC, pb.worker_category, pbl.worker_name_snapshot LIMIT %s"
    params.append(int(limit))

    c.execute(q, params)
    raw_rows = c.fetchall() or []

    excluded_missing_pay_date_count = 0
    if resolved_type == "monthly_paid" and has_official_pay_date:
        # Informational: lines matching other filters but missing official_pay_date.
        miss_q = """
            SELECT COUNT(*) AS cnt
            FROM payout_batch_lines pbl
            INNER JOIN payout_batches pb ON pb.id = pbl.batch_id
            WHERE pb.organization_id = %s AND pbl.organization_id = %s
              AND pb.official_pay_date IS NULL
        """
        miss_params: list[Any] = [int(organization_id), int(organization_id)]
        if worker_category and worker_category != "all":
            miss_q += " AND pb.worker_category = %s"
            miss_params.append(worker_category)
        if user_id:
            miss_q += " AND pbl.user_id = %s"
            miss_params.append(int(user_id))
        if payment_status and payment_status != "all":
            miss_q += " AND pbl.payment_status = %s"
            miss_params.append(str(payment_status))
        c.execute(miss_q, miss_params)
        miss_row = c.fetchone() or {}
        excluded_missing_pay_date_count = int(miss_row.get("cnt") or 0)

    df = _parse_ymd(date_from)
    dt = _parse_ymd(date_to)
    use_custom_range = resolved_type == "custom_range" and bool(df and dt)
    if resolved_type == "custom_range" and not (df and dt):
        raise ValueError("custom_range requires date_from and date_to")

    rows: list[dict] = []
    seen_line_ids: set[Any] = set()
    for raw in raw_rows:
        line_id = raw.get("line_id")
        if line_id is not None and line_id in seen_line_ids:
            continue
        batch = {
            "id": raw.get("batch_id"),
            "batch_name": raw.get("batch_name"),
            "worker_category": raw.get("worker_category"),
            "pay_period_start": raw.get("pay_period_start"),
            "pay_period_end": raw.get("pay_period_end"),
            "status": raw.get("batch_status"),
            "payout_details_finalized_at": raw.get("payout_details_finalized_at"),
            "official_pay_date": raw.get("official_pay_date") if has_official_pay_date else None,
        }
        line = {
            "id": line_id,
            "user_id": raw.get("user_id"),
            "worker_name_snapshot": raw.get("worker_name_snapshot"),
            "approved_hours": raw.get("approved_hours"),
            "ot_hours": raw.get("ot_hours"),
            "rate": raw.get("rate"),
            "ot_rate": raw.get("ot_rate"),
            "gross_amount": raw.get("gross_amount"),
            "total_amount": raw.get("total_amount"),
            "gross_wages": raw.get("gross_wages"),
            "sick_pay_amount": raw.get("sick_pay_amount"),
            "bonus_tip_amount": raw.get("bonus_tip_amount"),
            "reimbursement_amount": raw.get("reimbursement_amount"),
            "adjustments": raw.get("adjustments"),
            "payment_status": raw.get("payment_status"),
            "net_pay": raw.get("net_pay"),
            "payout_details_json": raw.get("payout_details_json"),
        }
        row = build_report_row(batch, line, report_type=resolved_type)
        if resolved_type == "monthly_paid" and row.get("payment_status_key") == "unpaid":
            continue
        if use_custom_range and not _row_matches_date_range(row, df, dt, date_basis=basis):
            continue
        if payroll_status and payroll_status != "all":
            if row["payroll_status_key"] != payroll_status:
                continue
        if line_id is not None:
            seen_line_ids.add(line_id)
        rows.append(row)

    totals = _sum_totals(rows)
    summary = _build_summary(rows, totals)
    from backend.payroll_report_analytics import (
        build_report_analytics,
        normalize_trend_range,
    )

    mode = "month" if resolved_type == "monthly_paid" else "period"
    cmp_range = normalize_trend_range(
        trend_range if trend_range is not None else comparison_range,
        mode=mode,
    )
    filters = {
        "report_type": resolved_type,
        "date_basis": basis if resolved_type == "custom_range" else "",
        "period_starts": starts,
        "period_ends": ends,
        "date_from": date_from or "",
        "date_to": date_to or "",
        "month": month_i,
        "year": year_i,
        "all_history": resolved_type == "all_history",
        "user_id": int(user_id) if user_id else None,
        "worker_category": worker_category or "all",
        "payroll_status": payroll_status or "all",
        "payment_status": payment_status or "all",
        "date_match_rule": match_rule,
        "comparison_range": cmp_range,
        "trend_range": cmp_range,
        "compare_with": compare_with or "",
        "include_employee_detail": bool(include_employee_detail),
    }
    out: dict[str, Any] = {
        "rows": rows,
        "totals": totals,
        "summary": summary,
        "count": len(rows),
        "report_type": resolved_type,
        "date_basis": basis if resolved_type == "custom_range" else None,
        "report_heading": report_heading(filters),
        "filters": filters,
        "date_match_rule": match_rule,
        "columns": [{"key": k, "label": lab} for k, lab in REPORT_COLUMNS],
        "categories": [
            {"value": "all", "label": "All categories"},
            *[{"value": c, "label": CATEGORY_LABELS[c]} for c in WORKER_CATEGORIES],
        ],
        "payroll_statuses": [
            {"value": "all", "label": "All payroll statuses"},
            *[
                {"value": k, "label": v}
                for k, v in DISPLAY_STATUS_LABELS.items()
            ],
        ],
        "payment_statuses": [
            {"value": "all", "label": "All payment statuses"},
            {"value": "pending", "label": "Pending"},
            {"value": "approved_unpaid", "label": "Approved — unpaid"},
            {"value": "paid", "label": "Paid"},
        ],
        "batch_statuses": list(BATCH_STATUSES),
    }
    if resolved_type == "monthly_paid":
        out["excluded_missing_pay_date_count"] = excluded_missing_pay_date_count
    if include_analytics:
        analytics = build_report_analytics(
            conn,
            organization_id,
            detail_rows=rows,
            report_type=resolved_type,
            filters=filters,
            comparison_range=cmp_range,
            compare_with=compare_with,
            trend_range=cmp_range,
        )
        # Prefer analytics summary enrichment while keeping batch_count from detail.
        merged_summary = {
            **summary,
            **(analytics.get("summary") or {}),
            "batch_count": summary.get("batch_count", 0),
            "unique_employees": summary.get("unique_employees", 0),
        }
        out["summary"] = merged_summary
        out["analytics"] = analytics
        out["groups"] = analytics.get("groups") or []
    if not include_employee_detail:
        # Dashboard-only: keep category analytics; hide employee-identifiable rows.
        out["rows"] = []
        out["groups"] = []
        out["employee_detail_restricted"] = True
        out["employee_detail_message"] = (
            "You do not have permission to view employee payroll details."
        )
        if out.get("analytics"):
            out["analytics"]["groups"] = []
            out["analytics"]["employee_summaries_by_category"] = {}
            out["analytics"]["access"] = {
                **(out["analytics"].get("access") or {}),
                "can_view_employee_detail": False,
            }
    return json_safe(out)


def list_report_employees(conn, organization_id: int) -> list[dict]:
    """Distinct employees that appear on payout lines (for report filter)."""
    from backend.payroll_operations import ensure_payout_batches_tables

    ensure_payout_batches_tables(conn.cursor())
    c = conn.cursor(dictionary=True)
    c.execute(
        """
        SELECT DISTINCT pbl.user_id, pbl.worker_name_snapshot AS name
        FROM payout_batch_lines pbl
        INNER JOIN payout_batches pb ON pb.id = pbl.batch_id
        WHERE pb.organization_id = %s AND pbl.user_id IS NOT NULL
        ORDER BY pbl.worker_name_snapshot
        """,
        (int(organization_id),),
    )
    out = []
    seen = set()
    for row in c.fetchall() or []:
        uid = row.get("user_id")
        if uid in seen:
            continue
        seen.add(uid)
        out.append({"user_id": uid, "name": row.get("name") or f"User {uid}"})
    return out


def list_report_periods(conn, organization_id: int) -> list[dict]:
    """Distinct pay periods whose batches are all paid/closed/finalized.

    Open, draft, or partially processed periods are omitted so pickers and
    dashboards never surface incomplete payroll weeks. Category mix is
    irrelevant (W-2-only / Temp-only weeks are included when complete).
    """
    from backend.payroll_operations import ensure_payout_batches_tables
    from backend.payroll_report_analytics import list_org_periods_asc

    ensure_payout_batches_tables(conn.cursor())
    periods = list_org_periods_asc(conn, organization_id, require_complete=True)
    return [
        {
            "pay_period_start": ps,
            "pay_period_end": pe,
            "label": f"{ps} – {pe}",
        }
        for ps, pe in reversed(periods)
    ]


def report_heading(filters: dict) -> str:
    """Dynamic title for screen, PDF, and Excel metadata."""
    rt = str(filters.get("report_type") or "").strip().lower()
    if rt == "all_history" or filters.get("all_history"):
        return "Payroll Reports — All History"
    if rt == "monthly_paid" and filters.get("month") and filters.get("year"):
        m = int(filters["month"])
        y = int(filters["year"])
        return f"Monthly Payroll Paid: {calendar.month_name[m]} {y}"
    if filters.get("period_starts") and filters.get("period_ends"):
        pairs = list(zip(filters["period_starts"], filters["period_ends"]))
        if len(pairs) == 1:
            return f"Payroll Period: {pairs[0][0]} – {pairs[0][1]}"
        return f"Payroll Period Report ({len(pairs)} periods)"
    if filters.get("date_from") and filters.get("date_to"):
        basis = str(filters.get("date_basis") or "pay_date")
        basis_label = "Pay Date" if basis == "pay_date" else "Payroll Period Overlap"
        return (
            f"Payroll Report: {filters['date_from']} – {filters['date_to']} "
            f"({basis_label} Basis)"
        )
    return "Payroll Reports"


def _filter_summary_text(filters: dict) -> str:
    parts = [report_heading(filters)]
    cat = filters.get("worker_category") or "all"
    if cat != "all":
        parts.append(CATEGORY_LABELS.get(cat, cat))
    if filters.get("payroll_status") and filters["payroll_status"] != "all":
        parts.append(
            DISPLAY_STATUS_LABELS.get(filters["payroll_status"], filters["payroll_status"])
        )
    if filters.get("payment_status") and filters["payment_status"] != "all":
        parts.append(_payment_status_label(filters["payment_status"]))
    return " · ".join(parts) if parts else "All records"


def build_payroll_report_xlsx(report: dict) -> bytes:
    """Excel workbook with real numeric/date cells (not formatted text)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, numbers

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Reports"
    filters = report.get("filters") or {}
    heading = report.get("report_heading") or report_heading(filters)
    ws.append([heading])
    ws.append([_filter_summary_text(filters)])
    ws.append([report.get("date_match_rule") or DATE_MATCH_RULE])
    summary = report.get("summary") or {}
    if summary:
        ws.append(
            [
                f"Periods: {summary.get('payroll_period_count', 0)}",
                f"Pay dates: {summary.get('official_pay_date_count', 0)}",
                f"Workers: {summary.get('unique_employees', 0)}",
                f"Gross: {round(_money(summary.get('gross_pay')), 2)}",
                f"EE taxes: {round(_money(summary.get('employee_tax_deductions')), 2)}",
                f"Net: {round(_money(summary.get('net_pay')), 2)}",
                f"ER taxes: {round(_money(summary.get('employer_taxes')), 2)}",
                f"Total cost: {round(_money(summary.get('total_payroll_cost')), 2)}",
            ]
        )
    else:
        ws.append([])

    # Detail rows ordered Period → Pay Date → Employee (via groups when present)
    excel_rows: list[dict] = []
    nested = report.get("groups") or []
    if nested:
        for period in nested:
            for pd in period.get("pay_dates") or []:
                for row in pd.get("rows") or []:
                    enriched = dict(row)
                    enriched["_group_period"] = period.get("payroll_period") or ""
                    enriched["_group_pay_date"] = pd.get("pay_date") or ""
                    excel_rows.append(enriched)
    else:
        excel_rows = list(report.get("rows") or [])

    # Insert grouping columns at front of export when nested groups exist
    group_cols = (
        (("_group_period", "Payroll Period"), ("_group_pay_date", "Pay Date Group"))
        if nested
        else ()
    )
    export_columns = (*group_cols, *REPORT_COLUMNS)

    headers = [lab for _, lab in export_columns]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    money_keys = set(MONEY_TOTAL_KEYS)
    hour_keys = set(HOUR_TOTAL_KEYS)
    date_keys = set(DATE_CELL_KEYS)
    header_row = ws.max_row
    data_start = header_row + 1

    for row in excel_rows:
        excel_vals = []
        for key, _ in export_columns:
            val = row.get(key)
            if key in date_keys:
                d = _parse_ymd(val)
                excel_vals.append(d if d else None)
            elif key in money_keys or key in hour_keys:
                excel_vals.append(round(_money(val), 2))
            elif key == "batch_id":
                excel_vals.append(int(val) if val is not None and str(val).strip() != "" else None)
            else:
                excel_vals.append(val if val is not None else "")
        ws.append(excel_vals)
        r = ws.max_row
        for col_idx, (key, _) in enumerate(export_columns, start=1):
            cell = ws.cell(row=r, column=col_idx)
            if key in date_keys and cell.value is not None:
                cell.number_format = "YYYY-MM-DD"
            elif key in money_keys:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            elif key in hour_keys:
                cell.number_format = "0.00"

    totals = report.get("totals") or {}
    total_row = []
    first_label_done = False
    for key, _lab in export_columns:
        if key in totals:
            total_row.append(round(_money(totals[key]), 2))
        elif not first_label_done:
            total_row.append("Totals")
            first_label_done = True
        else:
            total_row.append("")
    ws.append(total_row)
    totals_excel_row = ws.max_row
    for cell in ws[totals_excel_row]:
        cell.font = Font(bold=True)
    for col_idx, (key, _) in enumerate(export_columns, start=1):
        cell = ws.cell(row=totals_excel_row, column=col_idx)
        if key in money_keys:
            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        elif key in hour_keys:
            cell.number_format = "0.00"

    note = totals_excel_row + 2
    ws.cell(
        row=note,
        column=1,
        value=(
            "OT Premium is only the additional amount above the regular hourly rate. "
            "Regular/Base Earnings include overtime hours at the regular rate. "
            "Regular/Base + OT Premium + Other Earnings = Gross Pay. "
            "Total Payroll Cost = Gross Pay + Employer Taxes (stored only). "
            "Pay Date is the batch official_pay_date (no period-end fallback). "
            "Monthly Paid detail is ordered Payroll Period → Pay Date → Employee."
        ),
    )
    ws.freeze_panes = f"A{data_start}"
    for col_idx, (key, _) in enumerate(export_columns, start=1):
        letter = ws.cell(row=header_row, column=col_idx).column_letter
        if key in date_keys or key in ("pay_period_start", "pay_period_end", "pay_date", "finalized_date"):
            ws.column_dimensions[letter].width = 14
        elif key in ("employee_name", "_group_period"):
            ws.column_dimensions[letter].width = 22
        elif key in ("employee_category", "batch_name", "payroll_status", "payment_status", "_group_pay_date"):
            ws.column_dimensions[letter].width = 16
        elif key in money_keys or key in hour_keys:
            ws.column_dimensions[letter].width = 12

    # Analytics sheets
    analytics = report.get("analytics") or {}
    categories = analytics.get("category_breakdown") or []
    workforce_totals = analytics.get("workforce_totals") or {}
    if categories:
        ws_wf = wb.create_sheet("Workforce Breakdown")
        ws_wf.append([heading, "Workforce breakdown"])
        ws_wf.append(
            [
                "Category",
                "Head Count",
                "Regular Hours",
                "OT Hours",
                "Regular/Base Earnings",
                "OT Premium",
                "Other Earnings",
                "Gross Payroll",
                "Employer Tax",
                "Total Payroll Cost",
                "Avg Employer Cost / Hour",
            ]
        )
        for cell in ws_wf[ws_wf.max_row]:
            cell.font = Font(bold=True)

        def _rate_cell(val):
            return round(_money(val), 2) if val is not None else None

        def _wf_values(c):
            return [
                c.get("label") or c.get("worker_category"),
                c.get("head_count") or c.get("worker_count") or 0,
                round(_money(c.get("regular_hours")), 2),
                round(_money(c.get("ot_hours")), 2),
                round(_money(c.get("base_earnings")), 2),
                round(_money(c.get("ot_premium")), 2),
                round(_money(c.get("other_earnings")), 2),
                round(_money(c.get("gross_pay")), 2),
                round(_money(c.get("employer_taxes")), 2),
                round(_money(c.get("total_payroll_cost")), 2),
                _rate_cell(c.get("avg_cost_per_hour")),
            ]

        for c in categories:
            ws_wf.append(_wf_values(c))
        ws_wf.append(_wf_values({**workforce_totals, "label": "Total"}))
        for cell in ws_wf[ws_wf.max_row]:
            cell.font = Font(bold=True)

    mc = analytics.get("month_comparison") or []
    pc = analytics.get("period_comparison") or []
    mode = analytics.get("comparison_mode") or ("month" if mc else "period")
    compare_with = analytics.get("compare_with") or filters.get("compare_with") or ""
    trend_n = analytics.get("trend_range") or analytics.get("comparison_range") or 4

    if mc and mode == "month":
        ws2 = wb.create_sheet("Month Comparison")
        ws2.append(
            [
                heading,
                f"Month comparison · compare_with={compare_with} · show_trend=last {trend_n} months",
            ]
        )
        headers2 = [
            "Month",
            "Pay Dates",
            "Distinct Head Count",
            "Regular Hours",
            "OT Hours",
            "Regular/Base Earnings",
            "OT Premium",
            "Gross",
            "Employer Tax",
            "Total Payroll Cost",
            "Avg Employer Cost / Hour",
            "Δ Total Cost",
            "% Δ Total Cost",
            "Δ Hours",
            "% Δ Hours",
        ]
        headers2[0] = "Month / Period"
        ws2.append(headers2)
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True)
        for e in mc:
            delta = (e.get("delta_from_previous") or {}).get("total_payroll_cost")
            pct = (e.get("pct_from_previous") or {}).get("total_payroll_cost")
            d_hrs = (e.get("delta_from_previous") or {}).get("total_hours")
            p_hrs = (e.get("pct_from_previous") or {}).get("total_hours")
            ws2.append(
                [
                    e.get("label") or e.get("month"),
                    e.get("pay_dates_label"),
                    e.get("worker_count") or e.get("head_count"),
                    round(_money(e.get("regular_hours")), 2),
                    round(_money(e.get("ot_hours")), 2),
                    round(_money(e.get("base_earnings")), 2),
                    round(_money(e.get("ot_premium")), 2),
                    round(_money(e.get("gross_pay")), 2),
                    round(_money(e.get("employer_taxes")), 2),
                    round(_money(e.get("total_payroll_cost")), 2),
                    round(_money(e.get("avg_cost_per_hour")), 2)
                    if e.get("avg_cost_per_hour") is not None
                    else None,
                    round(_money(delta), 2) if delta is not None else None,
                    round(_money(pct), 2) if pct is not None else None,
                    round(_money(d_hrs), 2) if d_hrs is not None else None,
                    round(_money(p_hrs), 2) if p_hrs is not None else None,
                ]
            )
            for cell in ws2[ws2.max_row]:
                cell.font = Font(bold=True)
            for per in e.get("periods") or []:
                ws2.append(
                    [
                        f"  {per.get('label') or per.get('payroll_period')}",
                        per.get("pay_dates_label"),
                        per.get("worker_count") or per.get("head_count"),
                        round(_money(per.get("regular_hours")), 2),
                        round(_money(per.get("ot_hours")), 2),
                        round(_money(per.get("base_earnings")), 2),
                        round(_money(per.get("ot_premium")), 2),
                        round(_money(per.get("gross_pay")), 2),
                        round(_money(per.get("employer_taxes")), 2),
                        round(_money(per.get("total_payroll_cost")), 2),
                        round(_money(per.get("avg_cost_per_hour")), 2)
                        if per.get("avg_cost_per_hour") is not None
                        else None,
                        None,
                        None,
                        None,
                        None,
                    ]
                )
    elif pc:
        ws2 = wb.create_sheet("Period Comparison")
        ws2.append(
            [
                heading,
                f"Period comparison · compare_with={compare_with} · show_trend=last {trend_n} periods",
            ]
        )
        headers2 = [
            "Payroll Period",
            "Pay Date(s)",
            "Head Count",
            "Regular Hours",
            "OT Hours",
            "Regular/Base Earnings",
            "OT Premium",
            "Gross",
            "Employer Tax",
            "Total Payroll Cost",
            "Avg Employer Cost / Hour",
            "Δ Total Cost",
            "% Δ Total Cost",
            "Δ Hours",
            "% Δ Hours",
        ]
        ws2.append(headers2)
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True)
        for e in pc:
            delta = (e.get("delta_from_previous") or {}).get("total_payroll_cost")
            pct = (e.get("pct_from_previous") or {}).get("total_payroll_cost")
            d_hrs = (e.get("delta_from_previous") or {}).get("total_hours")
            p_hrs = (e.get("pct_from_previous") or {}).get("total_hours")
            ws2.append(
                [
                    e.get("payroll_period"),
                    e.get("pay_dates_label"),
                    e.get("worker_count"),
                    round(_money(e.get("regular_hours")), 2),
                    round(_money(e.get("ot_hours")), 2),
                    round(_money(e.get("base_earnings")), 2),
                    round(_money(e.get("ot_premium")), 2),
                    round(_money(e.get("gross_pay")), 2),
                    round(_money(e.get("employer_taxes")), 2),
                    round(_money(e.get("total_payroll_cost")), 2),
                    round(_money(e.get("avg_cost_per_hour")), 2)
                    if e.get("avg_cost_per_hour") is not None
                    else None,
                    round(_money(delta), 2) if delta is not None else None,
                    round(_money(pct), 2) if pct is not None else None,
                    round(_money(d_hrs), 2) if d_hrs is not None else None,
                    round(_money(p_hrs), 2) if p_hrs is not None else None,
                ]
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_payroll_report_csv(report: dict) -> bytes:
    """CSV export of payroll detail (same columns as Excel detail sheet)."""
    import csv

    filters = report.get("filters") or {}
    nested = report.get("groups") or []
    excel_rows: list[dict] = []
    if nested:
        for period in nested:
            for pd in period.get("pay_dates") or []:
                for row in pd.get("rows") or []:
                    enriched = dict(row)
                    enriched["_group_period"] = period.get("payroll_period") or ""
                    enriched["_group_pay_date"] = pd.get("pay_date") or ""
                    excel_rows.append(enriched)
    else:
        excel_rows = list(report.get("rows") or [])

    group_cols = (
        (("_group_period", "Payroll Period"), ("_group_pay_date", "Pay Date Group"))
        if nested
        else ()
    )
    export_columns = (*group_cols, *REPORT_COLUMNS)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([lab for _, lab in export_columns])
    for row in excel_rows:
        out = []
        for key, _ in export_columns:
            val = row.get(key)
            if key in DATE_CELL_KEYS:
                d = _parse_ymd(val)
                out.append(d.isoformat() if d else "")
            elif key in MONEY_TOTAL_KEYS or key in HOUR_TOTAL_KEYS:
                out.append(round(_money(val), 2))
            elif val is None:
                out.append("")
            else:
                out.append(val)
        writer.writerow(out)
    # UTF-8 BOM helps Excel open CSV correctly
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _group_key_period(row: dict) -> str:
    return row.get("payroll_period") or (
        f"{row.get('pay_period_start') or ''} – {row.get('pay_period_end') or ''}".strip(" –")
        or "Unknown period"
    )


def _group_key_pay_date(row: dict) -> str:
    return row.get("pay_date") or row.get("official_pay_date") or "Pay Date Missing"


def _group_rows_for_pdf(report: dict) -> list[tuple[str, list[dict]]]:
    """Return ordered (group_heading, rows) based on report_type / date_basis.

    Monthly Payroll Paid (and default hierarchy): Payroll Period → Pay Date.
    Flat list of (heading, rows) for PDF sections; nested groups preferred when
    report['groups'] is present.
    """
    nested = report.get("groups")
    if nested:
        flat: list[tuple[str, list[dict]]] = []
        for period in nested:
            period_heading = period.get("heading") or (
                f"Payroll Period: {period.get('payroll_period') or ''}"
            )
            pay_dates = period.get("pay_dates") or []
            if not pay_dates:
                continue
            # First pay-date section carries the period heading as a marker;
            # build_payroll_report_html renders nested structure when groups exist.
            for pd in pay_dates:
                pay_heading = pd.get("heading") or f"Pay Date: {pd.get('pay_date')}"
                flat.append((f"{period_heading} · {pay_heading}", list(pd.get("rows") or [])))
        if flat:
            return flat

    rows = list(report.get("rows") or [])
    filters = report.get("filters") or {}
    rt = str(
        report.get("report_type") or filters.get("report_type") or ""
    ).strip().lower()
    basis = str(
        report.get("date_basis") or filters.get("date_basis") or "pay_date"
    ).strip().lower()

    if rt == "monthly_paid" or rt in ("payroll_period", "all_history") or (
        rt == "custom_range" and basis == "period_overlap"
    ):
        from backend.payroll_report_analytics import group_rows_by_period_then_pay_date

        nested_groups = group_rows_by_period_then_pay_date(rows)
        flat = []
        for period in nested_groups:
            for pd in period.get("pay_dates") or []:
                flat.append(
                    (
                        f"{period.get('heading')} · {pd.get('heading')}",
                        list(pd.get("rows") or []),
                    )
                )
        return flat

    if rt == "custom_range" and basis == "pay_date":
        by_pay = defaultdict(list)
        for r in rows:
            by_pay[_group_key_pay_date(r)].append(r)
        return [(f"Pay Date: {k}", by_pay[k]) for k in sorted(by_pay.keys())]

    by_period = defaultdict(list)
    for r in rows:
        by_period[_group_key_period(r)].append(r)
    return [(f"Payroll Period: {k}", by_period[k]) for k in sorted(by_period.keys())]


def build_payroll_report_html(report: dict) -> str:
    """HTML suitable for browser print / PDF download.

    Page 1: title/filters, KPI cards, charts, period comparison table.
    Following pages: Payroll Period → Pay Date detail groups.
    Landscape layout. Grand total at end.
    """
    from backend.payroll_report_analytics import build_analytics_chart_svgs

    filters = report.get("filters") or {}
    totals = report.get("totals") or {}
    summary = report.get("summary") or _build_summary(report.get("rows") or [], totals)
    analytics = report.get("analytics") or {}
    rt = str(report.get("report_type") or filters.get("report_type") or "").strip().lower()
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    def fmt_money(v: Any) -> str:
        return f"${_money(v):,.2f}"

    def fmt_hours(v: Any) -> str:
        return f"{_money(v):,.2f}"

    def fmt_pct(v: Any) -> str:
        if v is None:
            return "—"
        return f"{_money(v):,.1f}%"

    def fmt_delta(diff: Any, pct: Any) -> str:
        if diff is None:
            return "—"
        d = _money(diff)
        sign = "+" if d > 0 else ""
        base = f"{sign}{d:,.2f}"
        if pct is None:
            return base
        return f"{base} ({sign}{_money(pct):,.1f}%)"

    def summary_block(title: str, data: dict) -> str:
        return f"""
<div class="group-summary">
  <strong>{title}</strong>
  <span>Periods: {data.get('payroll_period_count', data.get('batch_count', 0))}</span>
  <span>Pay dates: {data.get('official_pay_date_count', 0)}</span>
  <span>Workers: {data.get('unique_employees', data.get('worker_count', 0))}</span>
  <span>Gross: {fmt_money(data.get('gross_pay'))}</span>
  <span>EE taxes: {fmt_money(data.get('employee_tax_deductions'))}</span>
  <span>Net: {fmt_money(data.get('net_pay'))}</span>
  <span>ER taxes: {fmt_money(data.get('employer_taxes'))}</span>
  <span>Total cost: {fmt_money(data.get('total_payroll_cost'))}</span>
</div>"""

    def rows_table(group_rows: list[dict], subtotal: dict) -> str:
        body = []
        for row in group_rows:
            cells = []
            for key, _ in PDF_COLUMNS:
                val = row.get(key)
                if key in MONEY_TOTAL_KEYS:
                    cells.append(f"<td class='num'>{fmt_money(val)}</td>")
                elif key in HOUR_TOTAL_KEYS:
                    cells.append(f"<td class='num'>{fmt_hours(val)}</td>")
                else:
                    cells.append(f"<td>{val or ''}</td>")
            body.append("<tr class='data-row'>" + "".join(cells) + "</tr>")

        sub_cells = []
        for key, _ in PDF_COLUMNS:
            if key == "employee_name":
                sub_cells.append("<td><strong>Subtotal</strong></td>")
            elif key in subtotal and key in (*HOUR_TOTAL_KEYS, *MONEY_TOTAL_KEYS):
                if key in HOUR_TOTAL_KEYS:
                    sub_cells.append(
                        f"<td class='num'><strong>{fmt_hours(subtotal[key])}</strong></td>"
                    )
                else:
                    sub_cells.append(
                        f"<td class='num'><strong>{fmt_money(subtotal[key])}</strong></td>"
                    )
            else:
                sub_cells.append("<td></td>")
        body.append("<tr class='subtotal'>" + "".join(sub_cells) + "</tr>")

        th = "".join(f"<th>{lab}</th>" for _, lab in PDF_COLUMNS)
        return f"""
<table>
<thead><tr>{th}</tr></thead>
<tbody>
{"".join(body)}
</tbody>
</table>"""

    heading = report.get("report_heading") or report_heading(filters)
    charts = build_analytics_chart_svgs(analytics) if analytics else {}
    kpis = analytics.get("kpis") or []
    ot = analytics.get("ot_summary") or {}
    categories = analytics.get("category_breakdown") or []
    workforce_totals = analytics.get("workforce_totals") or {}

    def fmt_kpi_val(kind: Any, val: Any) -> str:
        if val is None:
            return "—"
        if kind == "money":
            return fmt_money(val)
        if kind == "hours":
            return fmt_hours(val)
        return str(val)

    kpi_html = "".join(
        f"""<div class="kpi">
          <div class="kpi-label">{k.get('label')}</div>
          <div class="kpi-value">{fmt_kpi_val(k.get('kind'), k.get('current', k.get('value')))}</div>
          <div class="kpi-prev">Previous: {fmt_kpi_val(k.get('kind'), k.get('previous'))}</div>
          <div class="kpi-delta">{fmt_delta(k.get('diff'), k.get('pct'))}</div>
        </div>"""
        for k in kpis
    )
    ot_html = ""
    if ot:
        ot_html = f"""
<div class="ot-card">
  <div class="kpi-label">OT</div>
  <div class="kpi-value">{fmt_hours(ot.get('ot_hours') if ot.get('ot_hours') is not None else ot.get('value'))}</div>
  <div class="kpi-prev">{_money(ot.get('ot_pct_of_hours')):.2f}% of hours</div>
  <div class="kpi-prev">OT Premium {fmt_money(ot.get('ot_premium'))}</div>
</div>"""

    def _fmt_rate(val):
        return fmt_money(val) if val is not None else "—"

    def _wf_row_cells(c, *, strong=False):
        other = _money(c.get("other_earnings"))
        other_note = f" <span class='muted'>+other {fmt_money(other)}</span>" if abs(other) >= 0.005 else ""
        wrap = ("<strong>", "</strong>") if strong else ("", "")
        b, e = wrap

        def cell(val, *, money=False, hrs=False, rate=False):
            if rate:
                txt = _fmt_rate(val)
            elif money:
                txt = fmt_money(val)
            elif hrs:
                txt = fmt_hours(val)
            else:
                txt = str(val)
            return f"<td class='num'>{b}{txt}{e}</td>"

        return (
            f"<td>{b}{c.get('label') or c.get('worker_category') or ''}{e}</td>"
            + cell(c.get("head_count") or c.get("worker_count") or 0)
            + cell(c.get("regular_hours"), hrs=True)
            + cell(c.get("ot_hours"), hrs=True)
            + cell(c.get("base_earnings"), money=True)
            + cell(c.get("ot_premium"), money=True)
            + f"<td class='num'>{b}{fmt_money(c.get('gross_pay'))}{other_note}{e}</td>"
            + cell(c.get("employer_taxes"), money=True)
            + cell(c.get("total_payroll_cost"), money=True)
            + cell(c.get("avg_cost_per_hour"), rate=True)
        )

    wf_rows = []
    for c in categories:
        wf_rows.append(f"<tr>{_wf_row_cells(c)}</tr>")
    if categories:
        tot = {
            **workforce_totals,
            "label": "Total",
            "worker_category": "Total",
        }
        wf_rows.append(f"<tr class='subtotal'>{_wf_row_cells(tot, strong=True)}</tr>")
    workforce_table = f"""
<table class="cmp">
<thead><tr>
  <th>Category</th><th>HC</th><th>Regular Hrs</th><th>OT Hrs</th>
  <th>Regular/Base Earnings</th><th>OT Premium</th><th>Gross</th>
  <th>Employer Tax</th><th>Total Cost</th>
  <th>Avg Employer Cost</th>
</tr></thead>
<tbody>{"".join(wf_rows) if wf_rows else "<tr><td colspan='10'>No category data</td></tr>"}</tbody>
</table>"""

    mc = analytics.get("month_comparison") or []
    pc = analytics.get("period_comparison") or []
    cmp_mode = analytics.get("comparison_mode") or ("month" if mc else "period")
    cmp_rows = []
    if cmp_mode == "month" and mc:
        for e in mc:
            avg = e.get("avg_cost_per_hour")
            d_cost = (e.get("delta_from_previous") or {}).get("total_payroll_cost")
            p_cost = (e.get("pct_from_previous") or {}).get("total_payroll_cost")
            cmp_rows.append(
                "<tr class='month-row'>"
                f"<td><strong>{e.get('label') or e.get('month') or ''}</strong></td>"
                f"<td>{e.get('pay_dates_label') or '—'}</td>"
                f"<td class='num'>{e.get('worker_count', 0)}</td>"
                f"<td class='num'>{fmt_hours(e.get('total_hours'))}</td>"
                f"<td class='num'>{fmt_hours(e.get('ot_hours'))}</td>"
                f"<td class='num'>{fmt_money(e.get('gross_pay'))}</td>"
                f"<td class='num'>{fmt_money(e.get('total_payroll_cost'))}</td>"
                f"<td class='num'>{fmt_money(avg) if avg is not None else '—'}</td>"
                f"<td class='num'>{fmt_delta(d_cost, p_cost)}</td>"
                "</tr>"
            )
            for per in e.get("periods") or []:
                p_avg = per.get("avg_cost_per_hour")
                cmp_rows.append(
                    "<tr class='period-under-month'>"
                    f"<td style='padding-left:18px'>{per.get('label') or per.get('payroll_period') or ''}</td>"
                    f"<td>{per.get('pay_dates_label') or '—'}</td>"
                    f"<td class='num'>{per.get('worker_count', 0)}</td>"
                    f"<td class='num'>{fmt_hours(per.get('total_hours'))}</td>"
                    f"<td class='num'>{fmt_hours(per.get('ot_hours'))}</td>"
                    f"<td class='num'>{fmt_money(per.get('gross_pay'))}</td>"
                    f"<td class='num'>{fmt_money(per.get('total_payroll_cost'))}</td>"
                    f"<td class='num'>{fmt_money(p_avg) if p_avg is not None else '—'}</td>"
                    f"<td class='num'>—</td>"
                    "</tr>"
                )
        comparison_table = f"""
<table class="cmp">
<thead><tr>
  <th>Month / Period</th><th>Pay Date(s)</th><th>Workers</th><th>Hours</th><th>OT</th>
  <th>Gross</th><th>Total cost</th><th>$/hr</th><th>Δ cost</th>
</tr></thead>
<tbody>{"".join(cmp_rows) if cmp_rows else "<tr><td colspan='9'>No comparison months</td></tr>"}</tbody>
</table>"""
    else:
        for e in pc:
            avg = e.get("avg_cost_per_hour")
            d_cost = (e.get("delta_from_previous") or {}).get("total_payroll_cost")
            p_cost = (e.get("pct_from_previous") or {}).get("total_payroll_cost")
            cmp_rows.append(
                "<tr>"
                f"<td>{e.get('payroll_period') or ''}</td>"
                f"<td>{e.get('pay_dates_label') or '—'}</td>"
                f"<td class='num'>{e.get('worker_count', 0)}</td>"
                f"<td class='num'>{fmt_hours(e.get('total_hours'))}</td>"
                f"<td class='num'>{fmt_hours(e.get('ot_hours'))}</td>"
                f"<td class='num'>{fmt_money(e.get('base_earnings'))}</td>"
                f"<td class='num'>{fmt_money(e.get('ot_premium'))}</td>"
                f"<td class='num'>{fmt_money(e.get('gross_pay'))}</td>"
                f"<td class='num'>{fmt_money(e.get('employer_taxes'))}</td>"
                f"<td class='num'>{fmt_money(e.get('total_payroll_cost'))}</td>"
                f"<td class='num'>{fmt_money(avg) if avg is not None else '—'}</td>"
                f"<td class='num'>{fmt_delta(d_cost, p_cost)}</td>"
                "</tr>"
            )
        comparison_table = f"""
<table class="cmp">
<thead><tr>
  <th>Payroll Period</th><th>Pay Date(s)</th><th>Workers</th><th>Hours</th><th>OT</th>
  <th>Base</th><th>OT Prem</th><th>Gross</th><th>ER taxes</th><th>Total cost</th><th>$/hr</th><th>Δ cost</th>
</tr></thead>
<tbody>{"".join(cmp_rows) if cmp_rows else "<tr><td colspan='12'>No comparison periods</td></tr>"}</tbody>
</table>"""

    chart_grid = f"""
<div class="chart-grid">
  <div class="chart">{charts.get('cost_trajectory') or ''}</div>
  <div class="chart">{charts.get('hours_trajectory') or ''}</div>
  <div class="chart">{charts.get('employment_mix') or ''}</div>
  <div class="chart">{charts.get('cost_per_hour') or ''}</div>
</div>"""

    # Nested detail: prefer analytics groups
    nested = report.get("groups") or (analytics.get("groups") if analytics else None) or []
    group_html = []
    if nested:
        for period in nested:
            period_heading = period.get("heading") or f"Payroll Period: {period.get('payroll_period')}"
            period_sections = []
            for pd in period.get("pay_dates") or []:
                g_rows = list(pd.get("rows") or [])
                g_totals = pd.get("totals") or _sum_totals(g_rows)
                g_summary = pd.get("summary") or _build_summary(g_rows, g_totals)
                period_sections.append(
                    f"""
  <div class="pay-date-block">
    <h3>{pd.get('heading') or ('Pay Date: ' + str(pd.get('pay_date') or ''))}</h3>
    {summary_block("Pay-date summary", g_summary)}
    {rows_table(g_rows, g_totals)}
  </div>"""
                )
            if not period_sections:
                continue
            p_summary = period.get("summary") or {}
            group_html.append(
                f"""
<section class="group pdf-capture-page period-group">
  <h2>{period_heading}</h2>
  {summary_block("Period summary", p_summary)}
  {"".join(period_sections)}
</section>"""
            )
    else:
        for g_heading, g_rows in _group_rows_for_pdf(report):
            g_totals = _sum_totals(g_rows)
            g_summary = _build_summary(g_rows, g_totals)
            group_html.append(
                f"""
<section class="group pdf-capture-page">
  <h2>{g_heading}</h2>
  {summary_block("Group summary", g_summary)}
  {rows_table(g_rows, g_totals)}
</section>"""
            )

    grand_cells = []
    for key, _ in PDF_COLUMNS:
        if key == "employee_name":
            grand_cells.append("<td><strong>Grand Total</strong></td>")
        elif key in totals and key in (*HOUR_TOTAL_KEYS, *MONEY_TOTAL_KEYS):
            if key in HOUR_TOTAL_KEYS:
                grand_cells.append(
                    f"<td class='num'><strong>{fmt_hours(totals[key])}</strong></td>"
                )
            else:
                grand_cells.append(
                    f"<td class='num'><strong>{fmt_money(totals[key])}</strong></td>"
                )
        else:
            grand_cells.append("<td></td>")
    grand_th = "".join(f"<th>{lab}</th>" for _, lab in PDF_COLUMNS)
    excluded = ""
    if rt == "monthly_paid" and report.get("excluded_missing_pay_date_count"):
        excluded = (
            f"<p class='warn'>{int(report['excluded_missing_pay_date_count'])} "
            "line(s) excluded — Pay Date Missing.</p>"
        )

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>{heading}</title>
<style>
  @page {{ size: letter landscape; margin: 0.4in; }}
  body {{ font-family: "Segoe UI", system-ui, sans-serif; color: #0f172a; margin: 14px; font-size: 9.5px; }}
  .brand {{ display:flex; justify-content:space-between; align-items:flex-start; margin-bottom: 8px; }}
  .brand-mark {{ color:#0097b2; font-weight:800; font-size:1.1rem; letter-spacing:0.02em; }}
  h1 {{ color: #0097b2; font-size: 1.2rem; margin: 0 0 4px; }}
  h2 {{ color: #007a91; font-size: 1.02rem; margin: 12px 0 6px; page-break-after: avoid; }}
  h3 {{ color: #0f172a; font-size: 0.95rem; margin: 8px 0 4px; page-break-after: avoid; }}
  .meta {{ color: #475569; margin: 2px 0; }}
  .warn {{ color: #c2410c; font-weight: 600; }}
  .note {{ color: #64748b; font-size: 8.5px; margin: 6px 0 10px; max-width: 1100px; }}
  .dashboard {{ page-break-after: always; break-after: page; }}
  .kpi-grid {{ display:grid; grid-template-columns: repeat(6, 1fr); gap: 6px; margin: 6px 0 8px; }}
  .kpi {{ background:#f8fafc; border:1px solid #e2e8f0; border-radius:6px; padding:6px 7px; }}
  .kpi-label {{ color:#64748b; font-size:8px; text-transform:uppercase; letter-spacing:0.03em; }}
  .kpi-value {{ font-weight:700; font-size:12px; color:#0f172a; margin-top:2px; }}
  .kpi-prev {{ color:#64748b; font-size:8px; margin-top:2px; }}
  .muted {{ color:#94a3b8; font-size:8px; }}
  .kpi-delta {{ color:#475569; font-size:8px; margin-top:2px; font-weight:600; }}
  .ot-card {{ background:#e6f5f8; border:1px solid #c5e7ee; border-radius:6px; padding:6px 8px; max-width:220px; margin: 4px 0 8px; }}
  .section-title {{ color:#007a91; font-size:11px; font-weight:700; margin: 8px 0 4px; }}
  .chart-grid {{ display:grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-bottom: 8px; }}
  .chart {{ border:1px solid #e2e8f0; border-radius:6px; padding:4px; background:#fff; }}
  .chart svg {{ max-width:100%; height:auto; display:block; }}
  .group {{ page-break-inside: avoid; break-inside: avoid; margin-bottom: 12px; }}
  .period-group {{ page-break-before: auto; }}
  .pay-date-block {{ margin-top: 6px; page-break-inside: avoid; }}
  .group-summary {{ display: flex; flex-wrap: wrap; gap: 8px 14px; margin: 4px 0 8px;
    color: #334155; background: #f8fafc; padding: 6px 8px; border-radius: 4px; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 8px; }}
  table.cmp {{ font-size: 8.5px; }}
  thead {{ display: table-header-group; }}
  th, td {{ padding: 4px 5px; border-bottom: 1px solid #e2e8f0; text-align: left; vertical-align: top; }}
  th {{ background: #f8fafc; color: #007a91; white-space: nowrap; }}
  td.num, th.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  tr.data-row, tr.subtotal, tr.total {{ page-break-inside: avoid; break-inside: avoid; }}
  tr.subtotal {{ background: #f1f5f9; }}
  tr.total {{ background: #e2e8f0; font-weight: 700; }}
  .grand {{ margin-top: 16px; }}
  .footer-meta {{ color:#94a3b8; font-size:8px; margin-top: 8px; }}
</style></head><body>
<section class="dashboard pdf-capture-page" id="payroll-analytics-dashboard">
  <div class="brand">
    <div>
      <div class="brand-mark">VeeWash</div>
      <h1>{heading}</h1>
      <p class="meta">{_filter_summary_text(filters)}</p>
      <p class="meta">{report.get("date_match_rule") or DATE_MATCH_RULE}</p>
      {excluded}
    </div>
    <div class="meta" style="text-align:right">
      <div>Payroll Analytics</div>
      <div>Generated {generated}</div>
      <div>Comparison: {analytics.get('comparison_mode') or 'period'} · {analytics.get('compare_with') or ''} · show trend last {analytics.get('trend_range') or analytics.get('comparison_range') or filters.get('comparison_range') or 4}</div>
    </div>
  </div>
  {summary_block("Report summary", summary)}
  <div class="section-title">Executive Summary</div>
  <div class="kpi-grid">{kpi_html or "<div class='kpi'><div class='kpi-label'>No KPI data</div></div>"}</div>
  {ot_html}
  <div class="section-title">Workforce Breakdown</div>
  {workforce_table}
  <div class="section-title">Trends</div>
  {chart_grid}
  <div class="section-title">Period comparison</div>
  {comparison_table}
  <p class="note">OT Premium is only the additional amount above the regular hourly rate.
  Regular/Base + OT Premium + Other Earnings = Gross. Total Payroll Cost = Gross + Employer Taxes (EE taxes are not added again).
  Employment mix stacks W-2 gross + employer taxes with Temp and 1099 gross.
  Dashboard and detail use the same filtered dataset. Pay Date is official_pay_date only.</p>
</section>
{"".join(group_html)}
<section class="grand pdf-capture-page">
  <h2>Grand Total</h2>
  <table>
  <thead><tr>{grand_th}</tr></thead>
  <tbody>
  <tr class="total">{"".join(grand_cells)}</tr>
  </tbody>
  </table>
  <p class="footer-meta">VeeWash Payroll Reports · {heading} · Generated {generated}</p>
</section>
</body></html>"""
